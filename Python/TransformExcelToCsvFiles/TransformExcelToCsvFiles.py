# -*-coding:Utf-8 -*

import glob
import os

import pandas
from logger import logger_config

from openpyxl import load_workbook

import param


def transform_text_to_valid_excel_date(initial_text: str) -> str:
    transformed_text: str = initial_text.replace("UTC+1", "")
    transformed_text = transformed_text.replace("UTC+2", "")
    transformed_text = transformed_text.replace("à ", "")
    return transformed_text


def get_list_of_excel_files_names() -> list[str]:
    dir_path = r"*.xls*"
    res = glob.glob(dir_path)
    return res


def transform_excel_file_to_csv_files(excel_file_name_with_extension: str) -> None:
    excel_file_name_without_extension = excel_file_name_with_extension.split(".")[0]

    with logger_config.stopwatch_with_label(f"transform_excel_file_to_csv_files: {excel_file_name_with_extension}"):
        with logger_config.stopwatch_with_label(
            "Load workbook:" + excel_file_name_with_extension, inform_beginning=True
        ):
            workbook = load_workbook(filename=excel_file_name_with_extension)

        workbook_sheetnames = workbook.sheetnames
        logger_config.print_and_log_info(
            "Workbook:" + excel_file_name_without_extension + " has following sheets:" + str(workbook_sheetnames)
        )

        if not os.path.exists(excel_file_name_without_extension):
            logger_config.print_and_log_info("Create folder " + excel_file_name_without_extension)
            os.makedirs(excel_file_name_without_extension)

        for sheet_name in workbook_sheetnames:
            with logger_config.stopwatch_with_label(
                f"transform_excel_file_to_csv_files: {excel_file_name_with_extension} sheet {sheet_name}"
            ):
                with logger_config.stopwatch_with_label(
                    f"read_excel {excel_file_name_with_extension} sheet {sheet_name}"
                ):
                    read_file = pandas.read_excel(excel_file_name_with_extension, sheet_name=sheet_name)
                output_csv_file = excel_file_name_without_extension + "/" + sheet_name + ".csv"
                with logger_config.stopwatch_with_label("Extract sheet " + sheet_name + " to " + output_csv_file):
                    read_file.to_csv(output_csv_file, sep=param.csv_separator, index=None, header=True)


def transform_all_excel_files_to_csv_files() -> None:

    list_of_excel_files_names = get_list_of_excel_files_names()

    logger_config.print_and_log_info(
        "Number of excel files found:" + str(len(list_of_excel_files_names)) + " : " + str(list_of_excel_files_names)
    )

    for excel_file_name in list_of_excel_files_names:
        transform_excel_file_to_csv_files(excel_file_name)


if __name__ == "__main__":

    with logger_config.application_logger("TransformExcelToCsvFiles"):
        transform_all_excel_files_to_csv_files()
