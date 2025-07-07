# Standard

from warnings import deprecated

from enum import Enum, auto
import argparse
import fnmatch
import inspect
import logging
import os
import pickle
import queue
import shutil
import threading
import time
from contextlib import contextmanager
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set

import pywintypes


import openpyxl
import xlwings

from selenium.webdriver.remote.webdriver import WebDriver as RemoteWebDriver

import selenium.webdriver.chrome.options
import selenium.webdriver.firefox.options
from common import download_utils, file_utils

# Other libraries
from logger import logger_config

# Third Party
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chromium.webdriver import ChromiumDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select, WebDriverWait

DEFAULT_DOWNLOAD_DIRECTORY = os.path.expandvars(r"%userprofile%\downloads")


DML_FILE_DOWNLOADED_PATTERN = "DML_NEXTEO_ATS+_V*.xlsm"
DML_FILE_WITHOUT_USELESS_SHEETS_PATH = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14_without_useless_sheets.xlsm"
DML_FILE_WITHOUT_USELESS_COLUMNS = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14_without_useless_columns.xlsm"
DML_FILE_WITH_USELESS_COLUMNS_CLEANED = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14_with_useless_columns_cleaned.xlsm"
DML_FILE_WITHOUT_FORMULA_REPLACED_BY_VALUE = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14_without_formula.xlsm"
DML_FILE_WITHOUT_LINKS = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14_without_links.xlsm"

DML_RAW_DOWNLOADED_FROM_RHAPSODY_FILE_PATH = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14_raw_from_rhapsody.xlsm"
DML_FILE_FINAL_DESTINATION_PATH = f"{DEFAULT_DOWNLOAD_DIRECTORY}\\DML_NEXTEO_ATS+_V14.xlsm"


USEFUL_DML_SHEET_NAME = "Database"
ALLOWED_DML_SHEETS_NAMES = [USEFUL_DML_SHEET_NAME]
EXCEL_INTERNAL_RESERVED_SHEETS_NAMES = ["Register"]

DOWNLOADED_FILES_FINAL_DIRECTORY = "Input"
OUTPUT_PARENT_DIRECTORY_DEFAULT_NAME = "output_save_cfx_request_results"

EXCEL_FILE_EXTENSION = ".xlsx"

FIRST_LINE_TO_REMOVE_RANGES = ["1:1"]
# RANGES_TO_REMOVE = ["A:G", "L:P", "S:W", "AA:AC"]

COLUMNS_NAMES_TO_REMOVE = [
    "Unique ID",
    "Code Dico & Code Arborescence",
    "MemoEasy",
    "Niveau Arborescence",
    "Code Arborescence",
    "Nom Arborescence",
    "Code Dico",
    "Type Document",
    "Tranche",
    "Phase",
    "Désignation Phase",
    "Work Package",
    "WorkPackageLeader",
    "Jalon Contractuel",
    "Date Jalon Contractuel",
    "Commentaires Internes",
    "Commentaires MOE",
    "Référentiel",
    "Jalon GUIDE",
    "Work Product",
    "ID Macrotache",
    "% Macro Tache",
    "Montant",
    "Activité Planning",
    "Date Planning",
    "Lien Doc-Planning",
    "Baseline Livraison",
    "Previous Forecast Livraison",
    "À confirmer Livraison",
    "Durée Livraison 1è Version & Acceptation",
    "GCONF",
    "Jalon Fourniture",
    "Nb de commentaires bloquants encore ouverts",
    "Document Supprimé",
    "Dernière Soumission ",
    "Document technique (Yes/No)",
    "Confidentialité Document",
    "Référence SNCF",
    "ATS+/Nexteo",
    "% Avancement DML",
    "Dernier statut Document",
    "Version avec dernier statut",
    "FILTERED OR NOT",
    "IS LAST VERSION",
    "IS FIRST VERSION",
    "IS FIRST ACCEPTATIONTBDLAST REFUSAL",
    "VERSION & REVISION",
    "CHROMATIC DISCRIMINANT",
    "Numeros  SNCF Mobilités \nn°10-5969 632 \nà\nn°10-5970-637",
    "Statut acceptation",
    "Date Acceptation",
    "Baseline FA",
]


DOWNLOAD_FROM_RHAPSODY_ENABLED = False
REMOVE_USELESS_SHEETS_ENABLED = False
REMOVE_LINKS_SHEETS_ENABLED = False
REMOVE_USELESS_COLUMNS_ENABLED = False
REPLACE_FORMULA_BY_VALUE_ENABLED = True


def save_and_close_workbook(workbook_dml: xlwings.Book | openpyxl.Workbook, file_path: str) -> str:

    with logger_config.stopwatch_with_label(label=f"Save:{file_path}"):
        workbook_dml.save(file_path)

    with logger_config.stopwatch_with_label(label="Close workbook"):
        workbook_dml.close()

    return file_path


@dataclass
class DownloadAndCleanDMLApplication:
    output_parent_directory_name: str = OUTPUT_PARENT_DIRECTORY_DEFAULT_NAME
    output_downloaded_files_final_directory_path: str = DOWNLOADED_FILES_FINAL_DIRECTORY
    web_browser_download_directory = DEFAULT_DOWNLOAD_DIRECTORY

    def __post_init__(self) -> None:
        self.number_of_exceptions_caught = 0

    def run(self) -> None:

        dml_file_path = self.download_dml_file()
        dml_file_path = self.remove_useless_tabs_with_xlwings(dml_file_path)
        dml_file_path = self.remove_excel_external_links_with_xlwings(dml_file_path)
        dml_file_path = self.remove_useless_columns_with_xlwings(dml_file_path)
        dml_file_path = self.replace_formulas_with_values_with_xlwings(dml_file_path)

    def remove_useless_tabs_with_xlwings(self, dml_file_path: str) -> str:
        file_to_create_path = DML_FILE_WITHOUT_USELESS_SHEETS_PATH

        if not REMOVE_USELESS_SHEETS_ENABLED:
            return file_to_create_path

        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            workbook_dml = xlwings.Book(dml_file_path)

        logger_config.print_and_log_info("set formulas calculations to manual to improve speed")

        workbook_dml.app.calculation = "manual"
        sheets_names = workbook_dml.sheet_names
        logger_config.print_and_log_info(f"{len(sheets_names)} Sheets found: {sheets_names}")

        for sheet_name in sheets_names:
            if sheet_name in ALLOWED_DML_SHEETS_NAMES:
                logger_config.print_and_log_info(f"Allowed sheet:{sheet_name}")
            elif sheet_name in EXCEL_INTERNAL_RESERVED_SHEETS_NAMES:
                logger_config.print_and_log_info(f"ignore Excel internal reserved sheet:{sheet_name}")
            else:
                with logger_config.stopwatch_with_label(label=f"Removing sheet:{sheet_name}", inform_beginning=True):
                    # Accéder à la feuille que l'on veut supprimer
                    sheet_to_remove = xlwings.sheets[sheet_name]
                    # Supprimer la feuille
                    try:
                        sheet_to_remove.delete()
                    except Exception as exc:
                        logger_config.print_and_log_exception(exc)
            pass

        return save_and_close_workbook(workbook_dml, file_to_create_path)

    def remove_excel_external_links_with_xlwings(self, dml_file_path: str) -> str:
        file_to_create_path = DML_FILE_WITHOUT_LINKS

        if not REMOVE_LINKS_SHEETS_ENABLED:
            return file_to_create_path

        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            workbook_dml = xlwings.Book(dml_file_path)

        logger_config.print_and_log_info("set formulas calculations to manual to improve speed")

        workbook_dml.app.calculation = "manual"
        external_links_sources = workbook_dml.api.LinkSources()

        logger_config.print_and_log_info(f"{len(external_links_sources)} links found: {external_links_sources}")

        for external_links_source_name in external_links_sources:
            with logger_config.stopwatch_with_label(label=f"Removing link:{external_links_source_name}"):
                workbook_dml.api.BreakLink(Name=external_links_source_name, Type=1)  # Type=1 pour les liaisons de type Excel

        return save_and_close_workbook(workbook_dml, file_to_create_path)

    def replace_formulas_with_values_with_xlwings(self, dml_file_path: str) -> str:
        file_to_create_path = DML_FILE_WITHOUT_FORMULA_REPLACED_BY_VALUE

        if not REPLACE_FORMULA_BY_VALUE_ENABLED:
            return file_to_create_path

        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            workbook_dml = xlwings.Book(dml_file_path)

        number_of_cells_updated = 0
        number_of_cells_not_updated = 0

        # Parcourir toutes les feuilles
        for sheet in workbook_dml.sheets:
            sheet: xlwings.Sheet = sheet
            number_of_cells_to_parse = len(sheet.used_range)
            with logger_config.stopwatch_with_label(label=f"Handle sheet:{sheet.name} with {number_of_cells_to_parse} cells to parse", inform_beginning=True):
                # Parcourir toutes les cellules dans la zone utilisée de la feuille
                for cell in sheet.used_range:

                    number_of_cells_treated = number_of_cells_updated + number_of_cells_not_updated
                    if (number_of_cells_treated) % 500 == 0:
                        logger_config.print_and_log_info(f"sheet:{sheet.name} : {number_of_cells_treated} cells treated {number_of_cells_treated/number_of_cells_to_parse*100:.2f} %")

                    # Si la cellule contient une formule
                    if cell.formula != "":
                        # Remplacer la formule par sa valeur actuelle
                        cell.value = cell.value
                        number_of_cells_updated += 1
                    else:
                        number_of_cells_not_updated += 1

        logger_config.print_and_log_info(f"{number_of_cells_updated} cells updateds and {number_of_cells_not_updated} not updated")

        return save_and_close_workbook(workbook_dml, file_to_create_path)

    def remove_useless_columns_with_xlwings(self, dml_file_path: str) -> str:
        file_to_create_path = DML_FILE_WITHOUT_USELESS_COLUMNS

        if not REMOVE_USELESS_COLUMNS_ENABLED:
            return file_to_create_path

        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            workbook_dml = xlwings.Book(dml_file_path)

        for range_to_remove in FIRST_LINE_TO_REMOVE_RANGES:
            with logger_config.stopwatch_with_label(label=f"Remove:{range_to_remove}", inform_beginning=True):
                xlwings.Range(range_to_remove).delete()

        sht = workbook_dml.sheets[USEFUL_DML_SHEET_NAME]

        for column_name_to_remove in COLUMNS_NAMES_TO_REMOVE:

            # Obtenir toutes les valeurs de la première ligne
            headers = sht.range("A1").expand("right").value
            logger_config.print_and_log_info(f"{len(headers)} headers:{headers}")

            # Trouver l'index de la colonne à supprimer
            logger_config.print_and_log_error(f"removing column '{column_name_to_remove}'")
            try:
                col_index = headers.index(column_name_to_remove) + 1  # Ajouter 1 car Excel utilise des index de base 1
                logger_config.print_and_log_info(f"index of column {column_name_to_remove} is {col_index}")
                sht.range((1, col_index), (sht.cells.last_cell.row, col_index)).delete()  # Supprimer la colonne

            except ValueError as e:
                logger_config.print_and_log_exception(e)
                logger_config.print_and_log_error(f"La colonne '{column_name_to_remove}' n'a pas été trouvée.")
                for range_to_remove in FIRST_LINE_TO_REMOVE_RANGES:
                    with logger_config.stopwatch_with_label(label=f"Remove:{range_to_remove}", inform_beginning=True):
                        xlwings.Range(range_to_remove).delete()

        return save_and_close_workbook(workbook_dml, file_to_create_path)

    def download_dml_file(self) -> Optional[str]:

        file_to_create_path = DML_RAW_DOWNLOADED_FROM_RHAPSODY_FILE_PATH

        if not DOWNLOAD_FROM_RHAPSODY_ENABLED:
            return file_to_create_path

        driver: RemoteWebDriver = webdriver.Firefox()

        dml_download_url = "https://rhapsody.siemens.net/livelink/livelink.exe?func=ll&objId=79329709&objAction=Download"

        driver.get(dml_download_url)

        time.sleep(0.5)

        wait = WebDriverWait(driver, 10)
        more_providers_button = wait.until(expected_conditions.element_to_be_clickable((By.ID, "moreProviders")))
        more_providers_button.click()

        # Wait until the Azure option is visible and then click it
        azure_option = wait.until(expected_conditions.visibility_of_element_located((By.XPATH, "//li[@class='secondary-menu-item authprovider-choice' and @data-authhandler='Azure']")))

        download_file_detector = download_utils.DownloadFileDetector(
            directory_path=self.web_browser_download_directory,
            filename_pattern=DML_FILE_DOWNLOADED_PATTERN,
            timeout_in_seconds=30,
        )

        azure_option.click()

        file_downloaded_path: Optional[str] = download_file_detector.monitor_download()
        if not file_downloaded_path:
            logger_config.print_and_log_error("No downloaded file found")
            return None

        logger_config.print_and_log_info(f"File downloaded : {file_downloaded_path}, will be moved to {DML_RAW_DOWNLOADED_FROM_RHAPSODY_FILE_PATH}")

        time.sleep(5)

        shutil.move(file_downloaded_path, file_to_create_path)

        time.sleep(5)

        driver.quit()
        return file_to_create_path

    @deprecated("Does not work")
    def not_working_replace_formulas_with_values_with_openpyxl(self, dml_file_path: str) -> str:
        # Load the workbook
        with logger_config.stopwatch_with_label(label=f"Open workbook_data_only:{dml_file_path}", inform_beginning=True):
            workbook_data_only = openpyxl.load_workbook(dml_file_path, data_only=True)

        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            # Read the workbook again to write values (data_only doesn't write back)
            writable_workbook = openpyxl.load_workbook(dml_file_path)

        sheet_name = USEFUL_DML_SHEET_NAME
        # Select the readable and writable sheets
        readable_sheet = workbook_data_only[sheet_name]
        writable_sheet = writable_workbook[sheet_name]

        number_of_cells_updated = 0
        number_of_cells_not_updated = 0

        with logger_config.stopwatch_with_label(label="Iterate and remove formula", inform_beginning=True):

            # Iterate over all cells in the sheet
            for row in readable_sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        # If the cell contains a formula, replace it with its value
                        writable_sheet[cell.coordinate].value = cell.value
                        number_of_cells_updated += 1
                    else:
                        number_of_cells_not_updated += 1

        logger_config.print_and_log_info(f"{number_of_cells_updated} cells updateds and {number_of_cells_not_updated} not updated")

        return save_and_close_workbook(writable_workbook, DML_FILE_WITHOUT_FORMULA_REPLACED_BY_VALUE)

    @deprecated("Does not work")
    def not_working_remove_useless_columns_with_openpyxl(self, dml_file_path: str) -> str:
        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            workbook_dml = openpyxl.load_workbook(dml_file_path)

        sheet = workbook_dml[USEFUL_DML_SHEET_NAME]

        columns_to_remove_letters = ["A", "E"]

        for columns_to_remove_letter in columns_to_remove_letters:

            # Convert the column letter to a number
            column_number = openpyxl.utils.column_index_from_string(columns_to_remove_letter)

            # Delete the column
            sheet.delete_cols(column_number)

        return save_and_close_workbook(workbook_dml, DML_FILE_WITHOUT_USELESS_COLUMNS)

    @deprecated("Does not work")
    def not_working_clean_useless_columns_with_openpyxl(self, dml_file_path: str) -> None:
        with logger_config.stopwatch_with_label(label=f"Open:{dml_file_path}", inform_beginning=True):
            workbook_dml = openpyxl.load_workbook(dml_file_path)

        sheet = workbook_dml[USEFUL_DML_SHEET_NAME]
        columns_to_clean_letters = ["B", "C"]

        number_of_cells_updated = 0

        # Iterate over each column specified
        for column_to_remove in columns_to_clean_letters:
            with logger_config.stopwatch_with_label(label=f"Remove column:{column_to_remove}", inform_beginning=True):
                # Iterate over each cell in the column (excluding the header)
                for cell in sheet[column_to_remove][2:]:  # Assuming the second row is header, start from the second
                    cell.value = None
                    number_of_cells_updated += 1

        save_and_close_workbook(workbook_dml, DML_FILE_WITH_USELESS_COLUMNS_CLEANED)

    @deprecated("Does not work")
    def not_working_remove_useless_tabs_with_openpyxl(self, dml_file_path: str) -> None:
        logger_config.print_and_log_info(f"Open:{dml_file_path}")
        workbook_dml = openpyxl.load_workbook(dml_file_path)
        sheets_names = workbook_dml.sheetnames
        logger_config.print_and_log_info(f"{len(sheets_names)} Sheets found: {sheets_names}")
        # print(workbook_dml.sheetnames)
        for sheet_name in sheets_names:
            if sheet_name in ALLOWED_DML_SHEETS_NAMES:
                logger_config.print_and_log_info(f"Allowed sheet:{sheet_name}")
            elif sheet_name in EXCEL_INTERNAL_RESERVED_SHEETS_NAMES:
                logger_config.print_and_log_info(f"ignore Excel internal reserved sheet:{sheet_name}")
            else:
                logger_config.print_and_log_info(f"Removing sheet:{sheet_name}")
                workbook_dml.remove(workbook_dml[sheet_name])
            pass

        save_and_close_workbook(workbook_dml, DML_FILE_WITHOUT_USELESS_SHEETS_PATH)


def main() -> None:
    """Main function"""

    with logger_config.application_logger("DownloadAndCleanDML"):

        application: DownloadAndCleanDMLApplication = DownloadAndCleanDMLApplication()
        application.run()


if __name__ == "__main__":
    main()
