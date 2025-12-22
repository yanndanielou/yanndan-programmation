# Standard
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from enum import Enum, auto
from typing import List, cast

import openpyxl
import xlwings

# Other libraries
from logger import logger_config
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch, gencache
import os


# import pywintypes
from xlsxwriter.utility import xl_cell_to_rowcol, xl_col_to_name
from xlwings.constants import DeleteShiftDirection

from common import file_name_utils, file_utils

EXCEL_INTERNAL_RESERVED_SHEETS_NAMES = ["Register"]

EXCEL_NA_NUMERIC_VALUE = -2146826246

EXCEL_FILE_EXTENSION = ".xlsx"

# cf    https://learn.microsoft.com/fr-fr/office/vba/api/Excel.XlFileFormat
EXCEL_8_XLS_FORMAT_XL_FILE_FORMAT_VALUE = 56
EXCEL_WORKBOOK_DEFAULT_XLSX_FORMAT_XL_FILE_FORMAT_VALUE = 51

XL_SAVE_CONFLICT_RESOLUTION_XL_LOCAL_SESSION_CHANGES = 2  # Les modifications de l'utilisateur local sont toujours acceptées.
XL_SAVE_CONFLICT_RESOLUTION_XL_OTHER_SESSION_CHANGES = 3  # Les modifications de l'utilisateur local sont toujours rejetées.
XL_SAVE_CONFLICT_RESOLUTION_XL_USER_RESOLUTION = 1  # Une boîte de dialogue demande à l'utilisateur de résoudre le conflit.


def convert_xlsx_file_to_xls_with_win32com_dispatch(xlsx_file_full_path: str) -> str:
    with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls: {xlsx_file_full_path}"):
        xls_output_file_path = xlsx_file_full_path[:-1]
        with logger_config.stopwatch_with_label("convert_xlsx_file_to_xls: open excel application"):
            xl = Dispatch("Excel.Application")

        xl.DisplayAlerts = False

        with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls: open {xlsx_file_full_path}"):
            wb = xl.Workbooks.Add(xlsx_file_full_path)

        wb.CheckCompatibility = False
        wb.DoNotPromptForConvert = True

        with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls: save {xls_output_file_path}"):
            wb.SaveAs(xls_output_file_path, FileFormat=EXCEL_8_XLS_FORMAT_XL_FILE_FORMAT_VALUE)
        with logger_config.stopwatch_with_label("convert_xlsx_file_to_xls: quit excel"):
            xl.Quit()
        return xls_output_file_path


def convert_excel_file_to_xlsx_with_win32com_dispatch(excel_file_full_path: str) -> str:
    with logger_config.stopwatch_with_label(f"convert_excel_file_to_xlsx_with_win32com_dispatch: {excel_file_full_path}"):

        xls_output_file_path = (
            file_name_utils.get_file_folder_from_full_path(excel_file_full_path) + "/" + file_name_utils.get_file_name_without_extension_from_full_path(excel_file_full_path) + ".xlsx"
        )
        with logger_config.stopwatch_with_label("convert_excel_file_to_xlsx_with_win32com_dispatch: open excel application"):
            xl = Dispatch("Excel.Application")

        xl.DisplayAlerts = False

        with logger_config.stopwatch_with_label(f"convert_excel_file_to_xlsx_with_win32com_dispatch: open {excel_file_full_path}"):
            wb = xl.Workbooks.Add(excel_file_full_path)

        wb.CheckCompatibility = False
        wb.DoNotPromptForConvert = True

        with logger_config.stopwatch_with_label(f"convert_excel_file_to_xlsx_with_win32com_dispatch: save {xls_output_file_path}"):
            wb.SaveAs(xls_output_file_path, FileFormat=EXCEL_WORKBOOK_DEFAULT_XLSX_FORMAT_XL_FILE_FORMAT_VALUE)
        with logger_config.stopwatch_with_label("convert_excel_file_to_xlsx_with_win32com_dispatch: quit excel"):
            xl.Quit()
        return xls_output_file_path


def convert_xlsx_file_to_xls_with_openpyxl(xlsx_file_full_path: str) -> str:
    xls_output_file_path = xlsx_file_full_path[:-1]
    with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls_with_openpyxl: {xlsx_file_full_path}", inform_beginning=True):

        with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls_with_openpyxl: open {xlsx_file_full_path}"):
            workbook = openpyxl.load_workbook(xlsx_file_full_path)
        with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls_with_openpyxl: save {xls_output_file_path}"):
            workbook.save(xls_output_file_path)
        with logger_config.stopwatch_with_label(f"convert_xlsx_file_to_xls_with_openpyxl: close {xls_output_file_path}"):
            workbook.close()

    return xls_output_file_path


def close_all_xlwings() -> None:
    logger_config.print_and_log_info("close_all_xlwings")
    app = xlwings.apps.add()
    wb = xlwings.Book()
    wb.close()
    app.quit()


class XlWingOperationBase(ABC):

    @abstractmethod
    def do(self, workbook_dml: xlwings.Book) -> None:
        pass


@dataclass
class XlWingsOpenWorkbookOperation:
    input_excel_file_path: str
    excel_visibility: bool
    work_on_temp_copy_of_input_file: bool = False

    def __post_init__(self) -> None:
        self.temp_copy_file_path = ""
        self.temp_copy_directory_path = ""

    def do(self) -> xlwings.Book:
        if self.work_on_temp_copy_of_input_file:
            self.temp_copy_directory_path, self.temp_copy_file_path = file_utils.get_temporary_copy_of_file(self.input_excel_file_path)
            return XlWingsOpenWorkbookOperation.open_workbook(self.temp_copy_file_path, self.excel_visibility)
        else:
            return XlWingsOpenWorkbookOperation.open_workbook(self.input_excel_file_path, self.excel_visibility)

    @staticmethod
    def open_workbook(input_excel_file_path: str, excel_visibility: bool) -> xlwings.Book:
        with logger_config.stopwatch_with_label(label=f"Open {input_excel_file_path}", inform_beginning=True):
            xlwings.App(visible=excel_visibility)
            workbook_dml = xlwings.Book(input_excel_file_path)

        return workbook_dml


@dataclass
class XlWingsSaveWorkbookOperation(XlWingOperationBase):
    file_to_create_path: str

    def do(self, workbook_dml: xlwings.Book | openpyxl.Workbook) -> None:
        XlWingsSaveWorkbookOperation.save_workbook(workbook_dml, self.file_to_create_path)

    @staticmethod
    def save_workbook(workbook_dml: xlwings.Book | openpyxl.Workbook, file_path: str) -> str:
        with logger_config.stopwatch_with_label(f"save_workbook {file_path}", inform_beginning=True):
            success = False
            while not success:
                try:
                    with logger_config.stopwatch_with_label(label=f"Save:{file_path}"):
                        workbook_dml.save(file_path)

                    success = True
                except Exception as e:
                    logger_config.print_and_log_exception(e)
                    logger_config.print_and_log_error(f"Could not save:{file_path}, must be locked")
                    time.sleep(1)

            return file_path


@dataclass
class XlWingsSaveAndCloseWorkbookOperation(XlWingOperationBase):
    file_to_create_path: str

    def do(self, workbook_dml: xlwings.Book | openpyxl.Workbook) -> None:
        XlWingsSaveAndCloseWorkbookOperation.save_and_close_workbook(workbook_dml, self.file_to_create_path)

    @staticmethod
    def save_and_close_workbook(workbook_dml: xlwings.Book | openpyxl.Workbook, file_path: str) -> str:
        with logger_config.stopwatch_with_label(f"save_and_close_workbook {file_path}", inform_beginning=True):
            success = False
            while not success:
                try:
                    with logger_config.stopwatch_with_label(label=f"Save:{file_path}"):
                        workbook_dml.save(file_path)

                    success = True
                except Exception as e:
                    logger_config.print_and_log_exception(e)
                    logger_config.print_and_log_error(f"Could not save:{file_path}, must be locked")
                    time.sleep(1)

            success = False
            while not success:
                try:

                    with logger_config.stopwatch_with_label(label="Close workbook"):
                        workbook_dml.close()

                    success = True
                except Exception as e:
                    logger_config.print_and_log_exception(e)
                    logger_config.print_and_log_error(f"Could not close:{file_path}, must be locked")
                    time.sleep(1)

            return file_path


@dataclass
class XlWingsRemoveTabsOperation(XlWingOperationBase):
    sheets_to_keep_names: List[str]

    def do(self, workbook_dml: xlwings.Book) -> None:
        XlWingsRemoveTabsOperation.remove_tabs_with_xlwings(workbook_dml, self.sheets_to_keep_names)

    @staticmethod
    def remove_tabs_with_xlwings(workbook_dml: xlwings.Book, sheets_to_keep_names: List[str]) -> None:
        with logger_config.stopwatch_with_label(f"remove_tabs_with_xlwings , sheets_to_keep_names:{sheets_to_keep_names}", inform_beginning=True):

            sheets_names = workbook_dml.sheet_names
            number_of_initial_sheets_names = len(sheets_names)
            logger_config.print_and_log_info(f"{number_of_initial_sheets_names} Sheets found: {sheets_names}")

            for sheet_number, sheet_name in enumerate(sheets_names):
                if sheet_name in sheets_to_keep_names:
                    logger_config.print_and_log_info(f"Allowed sheet:{sheet_name}")
                elif sheet_name in EXCEL_INTERNAL_RESERVED_SHEETS_NAMES:
                    logger_config.print_and_log_info(f"ignore Excel internal reserved sheet:{sheet_name}")
                else:
                    with logger_config.stopwatch_with_label(
                        label=f"Removing {sheet_number+1}/{number_of_initial_sheets_names}th sheet:{sheet_name} : {round((sheet_number+1)/number_of_initial_sheets_names*100,2)}%",
                        inform_beginning=True,
                        monitor_ram_usage=True,
                    ):
                        try:
                            # Accéder à la feuille que l'on veut supprimer
                            sheet_to_remove = xlwings.sheets[sheet_name]
                            # Supprimer la feuille
                            sheet_to_remove.delete()
                        except Exception as exc:
                            logger_config.print_and_log_exception(type(exc))
                            logger_config.print_and_log_exception(exc)


@dataclass
class XlWingsSetExcelFormulasCalculationToManualOperation(XlWingOperationBase):

    def do(self, workbook_dml: xlwings.Book) -> None:
        XlWingsSetExcelFormulasCalculationToManualOperation.set_excel_formulas_calculation_to_manual_xlwings(workbook_dml)

    @staticmethod
    def set_excel_formulas_calculation_to_manual_xlwings(workbook_dml: xlwings.Book) -> None:
        with logger_config.stopwatch_with_label(f"set_excel_formulas_calculation_to_manual input_excel_file_path:{workbook_dml.name}", inform_beginning=True):
            logger_config.print_and_log_info("set formulas calculations to manual to improve speed")
            workbook_dml.app.calculation = "manual"


class XlWingsRemoveExcelExternalLinksOperation(XlWingOperationBase):

    def do(self, workbook_dml: xlwings.Book) -> None:
        XlWingsRemoveExcelExternalLinksOperation.remove_excel_external_links_with_xlwings(workbook_dml)

    @staticmethod
    def remove_excel_external_links_with_xlwings(workbook_dml: xlwings.Book) -> None:
        with logger_config.stopwatch_with_label(f"remove_excel_external_links_with_xlwings input_excel_file_path:{workbook_dml.name}", inform_beginning=True):
            external_links_sources = workbook_dml.api.LinkSources()

            if external_links_sources:
                logger_config.print_and_log_info(f"{len(external_links_sources)} links found: {external_links_sources}")

                for external_links_source_name in external_links_sources:
                    with logger_config.stopwatch_with_label(label=f"Removing link:{external_links_source_name}"):
                        workbook_dml.api.BreakLink(Name=external_links_source_name, Type=1)  # Type=1 pour les liaisons de type Excel
            else:
                logger_config.print_and_log_info("No external link found (pass)")


@dataclass
class XlWingsReplaceFormulasWithCurrentValueOperation(XlWingOperationBase):

    def do(self, workbook_dml: xlwings.Book) -> None:
        XlWingsReplaceFormulasWithCurrentValueOperation.replace_formulas_with_values_with_xlwings(workbook_dml)

    @staticmethod
    def replace_formulas_with_values_with_xlwings(workbook_dml: xlwings.Book) -> None:
        with logger_config.stopwatch_with_label(f"replace_formulas_with_values_with_xlwings input_excel_file_path:{workbook_dml.name}", inform_beginning=True):
            number_of_cells_updated = 0
            number_of_cells_not_updated = 0

            # Parcourir toutes les feuilles
            for sheet in workbook_dml.sheets:
                sheet = cast(xlwings.Sheet, sheet)
                number_of_cells_to_parse = len(sheet.used_range)
                with logger_config.stopwatch_with_label(label=f"Handle sheet:{sheet.name} with {number_of_cells_to_parse} cells to parse", inform_beginning=True):
                    # Parcourir toutes les cellules dans la zone utilisée de la feuille
                    for cell in sheet.used_range:

                        number_of_cells_treated = number_of_cells_updated + number_of_cells_not_updated
                        if (number_of_cells_treated) % 500 == 0:
                            logger_config.print_and_log_info(f"sheet:{sheet.name} : {number_of_cells_treated} cells treated {number_of_cells_treated/number_of_cells_to_parse*100:.2f} %")

                        # Si la cellule contient une formule
                        if cell.formula != "":
                            # Remplacer la formule par sa valeur actuelle
                            cell.value = cell.value
                            number_of_cells_updated += 1
                        else:
                            number_of_cells_not_updated += 1

            logger_config.print_and_log_info(f"{number_of_cells_updated} cells updateds and {number_of_cells_not_updated} not updated")


@dataclass
class XlWingsRemoveRangesOperation(XlWingOperationBase):
    ranges_to_remove: List[str] = field(default_factory=list)

    def do(self, workbook_dml: xlwings.Book) -> None:
        XlWingsRemoveRangesOperation.remove_ranges_with_xlwings(workbook_dml, self.ranges_to_remove)

    @staticmethod
    def remove_ranges_with_xlwings(workbook_dml: xlwings.Book, ranges_to_remove: List[str]) -> None:
        with logger_config.stopwatch_with_label(f"remove_ranges_with_xlwings input_excel_file_path:{workbook_dml.name}, ranges_to_remove:{ranges_to_remove}", inform_beginning=True):

            for range_to_remove in ranges_to_remove:
                with logger_config.stopwatch_with_label(label=f"Remove:{range_to_remove}", inform_beginning=True):
                    xlwings.Range(range_to_remove).delete()
        return


class ColumnRemovalOperationType(Enum):
    ALL_COLUMNS_AT_ONCE = auto()
    COLUMN_ONE_BY_ONE_USING_LETTER = auto()
    COLUMN_ONE_BY_ONE_USING_INDEX = auto()


@dataclass
class RemoveColumnsInstructions:
    input_excel_file_path: str
    sheet_name: str
    columns_to_remove_names: List[str]
    removal_operation_type: ColumnRemovalOperationType
    file_to_create_path: str
    assert_if_column_is_missing: bool


@dataclass
class XlWingsRemoveColumnsOperation(XlWingOperationBase):

    sheet_name: str
    columns_to_remove_names: List[str]
    assert_if_column_is_missing: bool
    removal_operation_type: ColumnRemovalOperationType = ColumnRemovalOperationType.COLUMN_ONE_BY_ONE_USING_LETTER

    def do(self, workbook_dml: xlwings.Book) -> None:
        XlWingsRemoveColumnsOperation.remove_columns_with_xlwings(
            workbook_dml=workbook_dml,
            sheet_name=self.sheet_name,
            columns_to_remove_names=self.columns_to_remove_names,
            assert_if_column_is_missing=self.assert_if_column_is_missing,
            removal_operation_type=self.removal_operation_type,
        )

    @staticmethod
    def remove_columns_with_xlwings(
        workbook_dml: xlwings.Book,
        sheet_name: str,
        columns_to_remove_names: List[str],
        assert_if_column_is_missing: bool,
        removal_operation_type: ColumnRemovalOperationType = ColumnRemovalOperationType.COLUMN_ONE_BY_ONE_USING_LETTER,
    ) -> None:

        number_of_columns_to_remove = len(columns_to_remove_names)

        with logger_config.stopwatch_with_label(
            f"remove_columns_with_xlwings input_excel_file_path:{workbook_dml.name}, sheet_name:{sheet_name}, {number_of_columns_to_remove} columns to remove: names:{columns_to_remove_names}",
            inform_beginning=True,
        ):

            sht: xlwings.Sheet = workbook_dml.sheets(sheet_name)

            # Obtenir toutes les valeurs de la première ligne
            at_beginning_headers: List[str] = sht.range("A1").expand("right").value
            logger_config.print_and_log_info(f"At beginning, {len(at_beginning_headers)} headers: {at_beginning_headers}")

            if removal_operation_type == ColumnRemovalOperationType.ALL_COLUMNS_AT_ONCE:

                all_col_initial_index_to_remove: List[int] = []
                for colum_it, column_name_to_remove in enumerate(columns_to_remove_names):
                    col_index_starting_0 = at_beginning_headers.index(column_name_to_remove)
                    all_col_initial_index_to_remove.append(col_index_starting_0)
                logger_config.print_and_log_info(f"all_col_initial_index_to_remove: {all_col_initial_index_to_remove}")

                all_col_initial_index_to_remove_sorted = list(reversed(sorted(all_col_initial_index_to_remove)))
                logger_config.print_and_log_info(f"all_col_initial_index_to_remove_sorted: {all_col_initial_index_to_remove_sorted}")

                all_columns_letters_to_remove = [xl_col_to_name(col_index) for col_index in all_col_initial_index_to_remove]
                logger_config.print_and_log_info(f"all_columns_letters_to_remove: {all_columns_letters_to_remove}")

                all_columns_letters_to_remove_range = ",".join([f"{letter}:{letter}" for letter in all_columns_letters_to_remove])
                logger_config.print_and_log_info(f"all_columns_letters_to_remove_range: {all_columns_letters_to_remove_range}")

                with logger_config.stopwatch_with_label(
                    label=f"Removing range {all_columns_letters_to_remove_range} columns {all_columns_letters_to_remove}",
                    inform_beginning=True,
                ):
                    sht.range(all_columns_letters_to_remove_range).api.Delete(DeleteShiftDirection.xlShiftToLeft)

            else:

                # Obtenir toutes les valeurs de la première ligne
                headers_found_in_excel: List[str] = sht.range("A1").expand("right").value
                logger_config.print_and_log_info(f"{len(headers_found_in_excel)} headers:{headers_found_in_excel}")

                if assert_if_column_is_missing:
                    # Check that all columns to remove are present
                    for column_to_remove_name in columns_to_remove_names:
                        assert column_to_remove_name in headers_found_in_excel, f"Column {column_to_remove_name} not found in excel among {headers_found_in_excel}"

                for colum_it, column_name_to_remove in enumerate(columns_to_remove_names):
                    # Obtenir toutes les valeurs de la première ligne
                    headers_found_in_excel = sht.range("A1").expand("right").value
                    logger_config.print_and_log_info(f"{len(headers_found_in_excel)} headers:{headers_found_in_excel}")

                    if column_name_to_remove in headers_found_in_excel:

                        # Trouver l'index de la colonne à supprimer
                        logger_config.print_and_log_info(
                            f"removing {colum_it+1}/{number_of_columns_to_remove}th column '{column_name_to_remove}' {round((colum_it+1)/number_of_columns_to_remove*100,2)}%"
                        )
                        col_index_starting_0 = headers_found_in_excel.index(column_name_to_remove)
                        col_letter = xl_col_to_name(col_index_starting_0)
                        logger_config.print_and_log_info(f"col_index:{col_index_starting_0}, col_letter:{col_letter}")
                        with logger_config.stopwatch_with_label(
                            label=f"Removing {colum_it+1}/{number_of_columns_to_remove}th column {column_name_to_remove} with index(starting0):{col_index_starting_0}, letter:{col_letter}.  {round((colum_it+1)/number_of_columns_to_remove*100,2)}%",
                            inform_beginning=False,
                        ):
                            if removal_operation_type == ColumnRemovalOperationType.COLUMN_ONE_BY_ONE_USING_LETTER:
                                col_range = f"{col_letter}:{col_letter}"

                                if sht.range(col_range).column_width == 0:
                                    logger_config.print_and_log_info(f"Attempting to delete hidden column {col_range}: show it")
                                    sht.range(col_range).column_width = 10
                                sht.range(col_range).delete()
                                # sht.range(col_range).delete(DeleteShiftDirection.xlShiftToLeft)
                            elif removal_operation_type == ColumnRemovalOperationType.COLUMN_ONE_BY_ONE_USING_INDEX:
                                col_index_starting_1 = headers_found_in_excel.index(column_name_to_remove) + 1
                                sht.range((1, col_index_starting_1), (sht.cells.last_cell.row, col_index_starting_1)).delete()  # Supprimer la colonne
                    else:
                        logger_config.print_and_log_error(f"Column {column_name_to_remove} not found in excel among {headers_found_in_excel}")

            at_end_headers: List[str] = sht.range("A1").expand("right").value
            logger_config.print_and_log_info(f"At the end, {len(at_end_headers)} headers: {at_end_headers}")


@dataclass
class XlWingsOperationsBatch:

    excel_visibility: bool
    input_excel_file_path: str
    file_to_create_path: str
    operations: List[XlWingOperationBase] = field(default_factory=list)

    def hide_excel(self) -> "XlWingsOperationsBatch":
        self.excel_visibility = False
        return self

    def show_excel(self) -> "XlWingsOperationsBatch":
        self.excel_visibility = True
        return self

    def add_operation(self, operation: XlWingOperationBase) -> "XlWingsOperationsBatch":
        self.operations.append(operation)
        return self

    def do(self) -> None:
        workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=self.input_excel_file_path, excel_visibility=self.excel_visibility).do()
        for operation in self.operations:
            operation.do(workbook_dml=workbook_dml)

        XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=self.file_to_create_path).do(workbook_dml)


def save_and_close_workbook(workbook_dml: xlwings.Book | openpyxl.Workbook, file_path: str) -> None:

    XlWingsSaveAndCloseWorkbookOperation(file_path).do(workbook_dml)


def remove_tabs_with_openpyxl(input_excel_file_path: str, file_to_create_path: str, sheets_to_keep_names: List[str]) -> str:
    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:

        with logger_config.stopwatch_with_label(
            f"remove_tabs_with_xlwings input_excel_file_path:{input_excel_file_path}, file_to_create_path:{file_to_create_path}, sheets_to_keep_names:{sheets_to_keep_names}", inform_beginning=True
        ):

            with logger_config.stopwatch_with_label(label=f"Open {temp_file_full_path}", inform_beginning=True):
                workbook = openpyxl.load_workbook(temp_file_full_path)

            sheets_names = workbook.sheetnames
            number_of_initial_sheets_names = len(sheets_names)
            logger_config.print_and_log_info(f"{number_of_initial_sheets_names} Sheets found: {sheets_names}")

            for sheet_number, sheet_name in enumerate(sheets_names):
                if sheet_name in sheets_to_keep_names:
                    logger_config.print_and_log_info(f"Allowed sheet:{sheet_name}")
                elif sheet_name in EXCEL_INTERNAL_RESERVED_SHEETS_NAMES:
                    logger_config.print_and_log_info(f"ignore Excel internal reserved sheet:{sheet_name}")
                else:
                    with logger_config.stopwatch_with_label(
                        label=f"Removing {sheet_number+1}/{number_of_initial_sheets_names}th sheet:{sheet_name} : {round((sheet_number+1)/number_of_initial_sheets_names*100,2)}%",
                        inform_beginning=True,
                        monitor_ram_usage=True,
                    ):
                        try:
                            # Accéder à la feuille que l'on veut supprimer
                            sheet_to_remove = workbook[sheet_name]
                            # Supprimer la feuille
                            workbook.remove(sheet_to_remove)

                            workbook.save(file_to_create_path + "_without_" + sheet_name)
                        except Exception as exc:
                            logger_config.print_and_log_exception(type(exc))
                            logger_config.print_and_log_exception(exc)

            workbook.save(file_to_create_path)
            workbook.close()
            return file_to_create_path


def remove_tabs_with_xlwings(input_excel_file_path: str, file_to_create_path: str, sheets_to_keep_names: List[str], excel_visibility: bool = False) -> str:
    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:
        with logger_config.stopwatch_with_label(
            f"remove_tabs_with_xlwings input_excel_file_path:{input_excel_file_path}, file_to_create_path:{file_to_create_path}, sheets_to_keep_names:{sheets_to_keep_names}", inform_beginning=True
        ):
            workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=temp_file_full_path, excel_visibility=excel_visibility).do()
            XlWingsRemoveTabsOperation(sheets_to_keep_names=sheets_to_keep_names).do(workbook_dml)
            XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=file_to_create_path).do(workbook_dml)
            return file_to_create_path


def set_excel_formulas_calculation_to_manual_xlwings(input_excel_file_path: str, file_to_create_path: str, excel_visibility: bool = False) -> str:
    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:

        with logger_config.stopwatch_with_label(
            f"set_excel_formulas_calculation_to_manual input_excel_file_path:{input_excel_file_path}, file_to_create_path:{file_to_create_path}", inform_beginning=True
        ):
            workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=temp_file_full_path, excel_visibility=excel_visibility).do()
            XlWingsSetExcelFormulasCalculationToManualOperation().do(workbook_dml)
            XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=file_to_create_path).do(workbook_dml)
            return file_to_create_path


def remove_excel_external_links_with_xlwings(input_excel_file_path: str, file_to_create_path: str, excel_visibility: bool = False) -> str:
    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:

        with logger_config.stopwatch_with_label(
            f"remove_excel_external_links_with_xlwings input_excel_file_path:{input_excel_file_path}, file_to_create_path:{file_to_create_path}", inform_beginning=True
        ):
            workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=temp_file_full_path, excel_visibility=excel_visibility).do()
            XlWingsRemoveExcelExternalLinksOperation().do(workbook_dml)
            XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=file_to_create_path).do(workbook_dml)
            return file_to_create_path


def replace_formulas_with_values_with_xlwings(input_excel_file_path: str, file_to_create_path: str, excel_visibility: bool = False) -> str:
    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:

        with logger_config.stopwatch_with_label(
            f"replace_formulas_with_values_with_xlwings input_excel_file_path:{input_excel_file_path}, file_to_create_path:{file_to_create_path}", inform_beginning=True
        ):
            workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=temp_file_full_path, excel_visibility=excel_visibility).do()
            XlWingsReplaceFormulasWithCurrentValueOperation().do(workbook_dml)
            XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=file_to_create_path).do(workbook_dml)
            return file_to_create_path


def remove_ranges_with_xlwings(input_excel_file_path: str, file_to_create_path: str, ranges_to_remove: List[str], excel_visibility: bool = False) -> str:
    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:

        with logger_config.stopwatch_with_label(
            f"remove_ranges_with_xlwings input_excel_file_path:{input_excel_file_path}, file_to_create_path:{file_to_create_path}, ranges_to_remove:{ranges_to_remove}", inform_beginning=True
        ):
            workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=temp_file_full_path, excel_visibility=excel_visibility).do()
            XlWingsRemoveRangesOperation(ranges_to_remove=ranges_to_remove).do(workbook_dml)
            XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=file_to_create_path).do(workbook_dml)
            return file_to_create_path


def remove_columns_with_openpyxl(
    remove_columns_instruction: RemoveColumnsInstructions,
) -> str:
    with file_utils.temporary_copy_of_file(remove_columns_instruction.input_excel_file_path) as temp_xlsm_file_full_path:
        with logger_config.stopwatch_with_label(
            f"remove_columns_with_xlwings input_excel_file_path:{remove_columns_instruction.input_excel_file_path}, sheet_name:{remove_columns_instruction.sheet_name}, file_to_create_path:{remove_columns_instruction.file_to_create_path}, columns_to_remove_names:{remove_columns_instruction.columns_to_remove_names}",
            inform_beginning=True,
        ):

            with logger_config.stopwatch_with_label(label=f"Open {temp_xlsm_file_full_path}", inform_beginning=True):
                workbook = openpyxl.load_workbook(temp_xlsm_file_full_path)

            worksheet = workbook[remove_columns_instruction.sheet_name]

            rows = worksheet.iter_rows(min_row=1, max_row=1)
            first_row = next(rows)
            at_beginning_headers = [c.value for c in first_row]

            logger_config.print_and_log_info(f"At beginning, {len(at_beginning_headers)} headers: {at_beginning_headers}")

            number_of_columns_to_remove = len(remove_columns_instruction.columns_to_remove_names)

            for colum_it, column_name_to_remove in enumerate(remove_columns_instruction.columns_to_remove_names):
                # Obtenir toutes les valeurs de la première ligne

                rows = worksheet.iter_rows(min_row=1, max_row=1)
                first_row = next(rows)
                headers_found_in_excel = [c.value for c in first_row]
                logger_config.print_and_log_info(f"{len(headers_found_in_excel)} headers:{headers_found_in_excel}")

                if column_name_to_remove in headers_found_in_excel:

                    # Trouver l'index de la colonne à supprimer
                    logger_config.print_and_log_info(f"removing {colum_it+1}/{number_of_columns_to_remove}th column '{column_name_to_remove}' {round((colum_it+1)/number_of_columns_to_remove*100,2)}%")

                    col_index_starting_0 = headers_found_in_excel.index(column_name_to_remove)
                    col_index_starting_1 = headers_found_in_excel.index(column_name_to_remove) + 1
                    col_letter = xl_col_to_name(col_index_starting_0)
                    logger_config.print_and_log_info(f"col_index:{col_index_starting_1}, col_letter:{col_letter}")
                    with logger_config.stopwatch_with_label(
                        label=f"Removing with xlwings {colum_it+1}/{number_of_columns_to_remove}th column {column_name_to_remove} with index(starting0):{col_index_starting_0}, letter:{col_letter}.  {round((colum_it+1)/number_of_columns_to_remove*100,2)}%",
                        inform_beginning=False,
                    ):
                        worksheet.delete_cols(col_index_starting_1)

                else:
                    logger_config.print_and_log_error(f"Column {column_name_to_remove} not found in excel among {headers_found_in_excel}")

            save_and_close_workbook(workbook, remove_columns_instruction.file_to_create_path)
            return remove_columns_instruction.file_to_create_path


def remove_columns_with_xlwings(
    remove_columns_instruction: RemoveColumnsInstructions,
    excel_visibility: bool = False,
) -> str:
    with file_utils.temporary_copy_of_file(remove_columns_instruction.input_excel_file_path) as temp_file_full_path:
        with logger_config.stopwatch_with_label(
            f"remove_columns_with_xlwings input_excel_file_path:{remove_columns_instruction.input_excel_file_path}, sheet_name:{remove_columns_instruction.sheet_name}, file_to_create_path:{remove_columns_instruction.file_to_create_path}, columns_to_remove_names:{remove_columns_instruction.columns_to_remove_names}",
            inform_beginning=True,
        ):
            workbook_dml = XlWingsOpenWorkbookOperation(input_excel_file_path=temp_file_full_path, excel_visibility=excel_visibility).do()
            XlWingsRemoveColumnsOperation(
                sheet_name=remove_columns_instruction.sheet_name,
                columns_to_remove_names=remove_columns_instruction.columns_to_remove_names,
                removal_operation_type=remove_columns_instruction.removal_operation_type,
                assert_if_column_is_missing=remove_columns_instruction.assert_if_column_is_missing,
            ).do(workbook_dml)
            XlWingsSaveAndCloseWorkbookOperation(file_to_create_path=remove_columns_instruction.file_to_create_path).do(workbook_dml)
            return remove_columns_instruction.file_to_create_path


def xl_column_name_to_index(column_name: str) -> int:
    """Index starts 0"""
    _, col = xl_cell_to_rowcol(column_name + "1")
    assert isinstance(col, int)
    return col


def copy_and_paste_excel_content_with_format_with_openpyxl(input_excel_file_path: str, sheet_name: str, output_excel_file_path: str) -> None:

    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:
        # Charger le fichier Excel source
        with logger_config.stopwatch_with_label(f"Open and load input {temp_file_full_path}", inform_beginning=True):
            wb_input = load_workbook(temp_file_full_path, data_only=True)  # `data_only=True` pour récupérer les valeurs, pas les formules
            if sheet_name not in wb_input.sheetnames:
                raise ValueError(f"La feuille '{sheet_name}' n'existe pas dans le fichier d'entrée.")

        sheet_input = wb_input[sheet_name]

        with logger_config.stopwatch_with_label("Créer un nouveau fichier Excel pour la sortie", inform_beginning=True):
            wb_output = Workbook()
            sheet_output = wb_output.active
            sheet_output.title = sheet_name

        with logger_config.stopwatch_with_label("Copier le contenu, les formats et le style cellule par cellule", inform_beginning=True):

            logger_config.print_and_log_info(f"{sum(1 for _ in sheet_input.iter_rows()) } rows")
            number_of_rows_processed = 0
            for row in sheet_input.iter_rows():
                if number_of_rows_processed % 100 == 0:
                    logger_config.print_and_log_info(f"Handling row {number_of_rows_processed}")
                for cell in row:
                    # Copier la valeur
                    new_cell = sheet_output.cell(row=cell.row, column=cell.column, value=cell.value)

                    # Copier les formats de la cellule
                    if cell.fill:
                        new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, fgColor=cell.fill.fgColor, bgColor=cell.fill.bgColor)
                    if cell.font:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color,
                        )
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent,
                        )
                    if cell.border:
                        new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                number_of_rows_processed += 1

        #
        with logger_config.stopwatch_with_label("Ajuster la largeur des colonnes et hauteur des lignes", inform_beginning=True):

            for col in sheet_input.columns:
                column_letter = get_column_letter(col[0].column)  # Récupère la lettre de la colonne
                sheet_output.column_dimensions[column_letter].width = sheet_input.column_dimensions[column_letter].width

            for row_idx, row in enumerate(sheet_input.iter_rows(), start=1):
                sheet_output.row_dimensions[row_idx].height = sheet_input.row_dimensions[row_idx].height

        with logger_config.stopwatch_with_label(f"Save output file {output_excel_file_path}"):
            wb_output.save(output_excel_file_path)
            print(f"Contenu copié avec succès dans : {output_excel_file_path}")


def copy_and_paste_excel_content_with_format_with_win32(
    input_excel_file_path: str, sheet_name: str, output_excel_file_path: str, replace_formulas_by_value: bool, excel_visibility: bool = False
) -> str:

    if os.path.exists(output_excel_file_path):
        logger_config.print_and_log_info(f"The file '{output_excel_file_path}' already exists. Deleting it...")
        os.remove(output_excel_file_path)  # Delete the existing file to avoid conflicts

    with file_utils.temporary_copy_of_file(input_excel_file_path) as temp_file_full_path:

        # Ensure the input file exists
        if not os.path.exists(temp_file_full_path):
            raise FileNotFoundError(f"The file '{temp_file_full_path}' does not exist.")

        with logger_config.stopwatch_with_label("Initialize Excel application (using COM)"):
            excel_app = gencache.EnsureDispatch("Excel.Application")
            excel_app.Visible = excel_visibility  # Make sure Excel doesn't open a UI window

        try:

            with logger_config.stopwatch_with_label(f"Open the input workbook {input_excel_file_path}", inform_beginning=True):
                wb_input = excel_app.Workbooks.Open(input_excel_file_path)

            with logger_config.stopwatch_with_label("Check if the sheet name exists in the input workbook"):
                try:
                    sheet_input = wb_input.Sheets(sheet_name)
                except Exception as exc:
                    raise ValueError(f"The sheet '{sheet_name}' was not found in the input file '{input_excel_file_path}'.") from exc

            with logger_config.stopwatch_with_label("Add a new workbook for the output"):
                wb_output = excel_app.Workbooks.Add()

            with logger_config.stopwatch_with_label("Copy the sheet from the input workbook to the new workbook", inform_beginning=True):
                sheet_input.Copy(Before=wb_output.Sheets(1))

            # Get the copied sheet (will always be the first sheet in the new workbook)
            sheet_copied = wb_output.Sheets(1)

            # logger_config.print_and_log_info(f"sheet_input.UsedRange:{sheet_input.UsedRange}")
            # logger_config.print_and_log_info(f"sheet_copied.UsedRange:{sheet_copied.UsedRange}")
            if replace_formulas_by_value:
                with logger_config.stopwatch_with_label("Remove formulas by pasting values only", inform_beginning=True):
                    # sheet_copied.UsedRange.Value = sheet_copied.UsedRange.Value
                    sheet_copied.UsedRange.Value = sheet_input.UsedRange.Value

            with logger_config.stopwatch_with_label(f"Save output workbook {output_excel_file_path}"):
                wb_output.SaveAs(
                    Filename=output_excel_file_path,
                    FileFormat=EXCEL_WORKBOOK_DEFAULT_XLSX_FORMAT_XL_FILE_FORMAT_VALUE,
                    ConflictResolution=XL_SAVE_CONFLICT_RESOLUTION_XL_LOCAL_SESSION_CHANGES,
                )

            with logger_config.stopwatch_with_label(f"Close output workbook {output_excel_file_path}"):
                wb_output.Close()  # Close the output workbook

            print(f"Excel sheet '{sheet_name}' was successfully copied and saved to '{output_excel_file_path}'.")
        finally:
            # Close the input workbook and quit the Excel application
            with logger_config.stopwatch_with_label(f"Close input workbook {temp_file_full_path}"):
                wb_input.Close(SaveChanges=False)

            with logger_config.stopwatch_with_label("Quit Excel app"):
                excel_app.Quit()

        return output_excel_file_path
