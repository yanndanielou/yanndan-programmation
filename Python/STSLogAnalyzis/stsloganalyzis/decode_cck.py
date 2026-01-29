import datetime
import os
import re
from collections import Counter
from dataclasses import dataclass, field
from enum import Enum, auto
from typing import Dict, List, Optional, Self, Tuple

import matplotlib.pyplot as plt
import plotly.graph_objects as go
from common import file_name_utils, string_utils, custom_iterator
from logger import logger_config
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# CONTENT_OF_FIELD_IN_CASE_OF_DECODING_ERROR = "!!! Decoding Error !!!"

LIAISON_PATTERN_STR = r".*(?P<liaison_full_name>Liaison (?P<liaison_id>\d+A?B?)).*"
LIAISON_PATTERN = re.compile(LIAISON_PATTERN_STR)

LINK_STATE_CHANGE_PATTERN_STR = r".*changement d'état : (?P<old_state>.*) => (?P<new_state>.*)"
LINK_STATE_CHANGE_PATTERN = re.compile(LINK_STATE_CHANGE_PATTERN_STR)


def save_cck_mpro_lines_in_excel(trace_lines: List["CckMproTraceLine"], output_folder_path: str, excel_output_file_name_without_extension: str) -> None:
    """
    Sauvegarde une liste de CckMproTraceLine dans un fichier Excel.

    Args:
        trace_lines: Liste des lignes de trace à sauvegarder
        excel_output_file_name: Nom du fichier Excel de sortie
    """
    if not trace_lines:
        logger_config.print_and_log_info("La liste des traces est vide. Aucun fichier créé.")
        return

    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CCK MPRO Traces"

    # Ajouter les en-têtes
    headers = ["Timestamp", "Date complète", "Liaison", "ID Liaison", "Ligne brute"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajouter les données
    for row_idx, trace_line in enumerate(trace_lines, start=2):
        ws.cell(row=row_idx, column=1).value = trace_line.decoded_timestamp
        ws.cell(row=row_idx, column=2).value = trace_line.raw_date_str
        ws.cell(row=row_idx, column=3).value = trace_line.liaison_full_name or "N/A"
        ws.cell(row=row_idx, column=4).value = trace_line.liaison_id or "N/A"
        ws.cell(row=row_idx, column=5).value = trace_line.full_raw_line.strip()

    # Ajuster la largeur des colonnes
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 80

    # Sauvegarder le fichier
    output_file_full_path = output_folder_path + "/" + excel_output_file_name_without_extension + file_name_utils.get_file_suffix_with_current_datetime() + ".xlsx"
    wb.save(output_file_full_path)
    logger_config.print_and_log_info(f"Fichier Excel créé: {output_file_full_path}")
    logger_config.print_and_log_info(f"Total de {len(trace_lines)} lignes sauvegardées")


def plot_bar_graph_list_cck_mpro_lines_by_period(trace_lines: List["CckMproTraceLine"], output_folder_path: str, label: str, interval_minutes: int = 10, do_show: bool = False) -> None:
    if not trace_lines:
        print("La liste des traces est vide.")
        return

    # Trier les trace_lines par timestamp
    trace_lines.sort(key=lambda x: x.decoded_timestamp)

    # Déterminer la période totale (du timestamp le plus tôt au plus tard)
    start_time = trace_lines[0].decoded_timestamp
    end_time = trace_lines[-1].decoded_timestamp

    # Créer des intervalles de temps
    interval_start_times: List[datetime.datetime] = []
    current_time = start_time
    while current_time <= end_time:
        interval_start_times.append(current_time)
        current_time += datetime.timedelta(minutes=interval_minutes)

    # Compter les éléments dans chaque intervalle
    interval_counts: Dict[Tuple[datetime.datetime, datetime.datetime], int] = Counter()
    for trace in trace_lines:
        for interval_start in interval_start_times:
            interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
            if interval_start <= trace.decoded_timestamp < interval_end:
                interval_counts[(interval_start, interval_end)] += 1
                break

    # Préparer les données pour le graphe
    x_labels = [f"{begin.strftime("%H:%M")} - {end.strftime("%H:%M")}" for begin, end in interval_counts.keys()]
    y_values = list(interval_counts.values())
    # Créer et exporter les données dans un fichier Excel
    excel_filename = f"interval_counts_{label.replace(' ', '_')}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Interval Counts"

    # Ajouter les en-têtes
    ws["A1"] = "Début Intervalle de temps"
    ws["B1"] = "Fin Intervalle de temps"
    ws["C1"] = "Nombre d'éléments"

    # Style des en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in [ws["A1"], ws["B1"], ws["C1"]]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajouter les données
    for idx, ((interval_begin, interval_end), count) in enumerate(interval_counts.items(), start=2):
        ws[f"A{idx}"] = interval_begin
        ws[f"B{idx}"] = interval_end
        ws[f"C{idx}"] = count
        ws[f"C{idx}"].alignment = Alignment(horizontal="center")

    # Ajuster la largeur des colonnes
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15

    # Ajouter un résumé
    summary_row = len(interval_counts) + 3
    ws[f"A{summary_row}"] = "Total"
    ws[f"B{summary_row}"] = len(trace_lines)
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"].alignment = Alignment(horizontal="center")

    # Sauvegarder le fichier
    wb.save(output_folder_path + "/" + excel_filename)
    logger_config.print_and_log_info(f"Fichier Excel créé: {excel_filename}")

    # Créer et sauvegarder le graphe en HTML avec Plotly
    html_filename = f"interval_counts_{label.replace(' ', '_')}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.html"
    fig = go.Figure(
        data=[
            go.Bar(
                x=x_labels,
                y=y_values,
                marker=dict(color="skyblue"),
                text=y_values,
                textposition="auto",
            )
        ]
    )

    fig.update_layout(
        title=f"{len(trace_lines)} {label} par périodes de {interval_minutes} minutes entre {start_time.strftime('%Y-%m-%d %H:%M')} et {end_time.strftime('%Y-%m-%d %H:%M')}",
        xaxis_title="Intervalles de temps (heure début - heure fin)",
        yaxis_title="Nombre de CckMproTraceLine",
        hovermode="x unified",
        template="plotly_white",
        height=600,
        width=1000,
    )

    fig.write_html(output_folder_path + "/" + html_filename)
    logger_config.print_and_log_info(f"Fichier HTML créé: {html_filename}")

    # Afficher le bar graph
    plt.figure(figsize=(10, 6))
    plt.bar(x_labels, y_values, color="skyblue", width=0.3)
    plt.xlabel("Intervalles de temps (heure début)")
    plt.ylabel("Nombre de CckMproTraceLine")
    plt.title(f"{len(trace_lines)} {label} par périodes de {interval_minutes} minutes entre {start_time.strftime("%Y-%m-%d %H:%M")} et {end_time.strftime("%Y-%m-%d %H:%M")}")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    if do_show:
        plt.show()


@dataclass
class CckMproTraceSpecificEvent:
    trace_line: "CckMproTraceLine"

    def __post_init__(self) -> None:
        assert self.trace_line.liaison
        self.liaison = self.trace_line.liaison


@dataclass
class CckMproProblemEnchainementNumeroProtocolaire(CckMproTraceSpecificEvent):
    additional_info: str = ""


class EtatLiaisonMpro(Enum):
    OK = auto()
    KO = auto()
    NON_CONNU = auto()


@dataclass
class CckMproTemporaryLossLink:
    loss_link_event: "CckMproChangementEtatLiaison"
    link_back_to_normal_event: "CckMproChangementEtatLiaison"

    def __post_init__(self) -> None:
        self.liaison = self.loss_link_event.liaison


@dataclass
class CckMproChangementEtatLiaison(CckMproTraceSpecificEvent):
    additional_info: str = ""

    def __post_init__(self) -> None:
        super().__post_init__()

        match = LINK_STATE_CHANGE_PATTERN.match(self.trace_line.full_raw_line)
        assert match
        self.previous_state = EtatLiaisonMpro[string_utils.text_to_valid_enum_value_text(match.group("old_state"))]
        self.new_state = EtatLiaisonMpro[match.group("new_state").upper()]
        # change_states = self.trace_line.full_raw_line.split("changement d'état : ")[1]
        # states = change_states.split(" => ")
        # self.previous_state = EtatLiaisonMpro[string_utils.text_to_valid_enum_value_text(states[0])]
        # self.new_state = EtatLiaisonMpro[string_utils.text_to_valid_enum_value_text(states[1])]


@dataclass
class CckMproTraceLibrary:
    name: str
    all_processed_lines: List["CckMproTraceLine"] = field(default_factory=list)
    all_processed_files: List["CckMproTraceFile"] = field(default_factory=list)
    all_problem_enchainement_numero_protocolaire: List["CckMproProblemEnchainementNumeroProtocolaire"] = field(default_factory=list)
    all_problem_enchainement_numero_protocolaire_per_link: Dict["CckMproLiaison", List["CckMproProblemEnchainementNumeroProtocolaire"]] = field(default_factory=dict)
    all_changement_etats_liaisons_mpro: List["CckMproChangementEtatLiaison"] = field(default_factory=list)
    all_changement_etats_liaisons_mpro_per_link: Dict["CckMproLiaison", List["CckMproChangementEtatLiaison"]] = field(default_factory=dict)
    lines_per_liaison: Dict[Optional["CckMproLiaison"], List["CckMproTraceLine"]] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.all_temporary_loss_link: List[CckMproTemporaryLossLink] = []
        self.all_liaisons: List[CckMproLiaison] = []
        self.liaison_by_identifier: Dict[str, CckMproLiaison] = {}

    def get_or_create_liaison(self, full_name: str, identifier: str) -> "CckMproLiaison":
        if identifier in self.liaison_by_identifier:
            return self.liaison_by_identifier[identifier]

        liaison = CckMproLiaison(full_name=full_name, identifier=identifier)
        self.liaison_by_identifier[identifier] = liaison
        self.all_liaisons.append(liaison)
        return liaison

    def load_folder(self, folder_full_path: str) -> Self:
        with logger_config.stopwatch_with_label("Load all files", inform_beginning=True, monitor_ram_usage=True):

            for dirpath, dirnames, filenames in os.walk(folder_full_path):
                for file_name in filenames:
                    cck_file = CckMproTraceFile(parent_folder_full_path=dirpath, file_name=file_name, library=self)
                    self.all_processed_files.append(cck_file)
                    self.all_processed_lines += cck_file.all_processed_lines
                    self.all_problem_enchainement_numero_protocolaire += cck_file.all_problem_enchainement_numero_protocolaire
                    self.all_changement_etats_liaisons_mpro += cck_file.all_changement_etats_liaisons_mpro

                    for key, value in cck_file.lines_per_liaison.items():
                        if key not in self.lines_per_liaison:
                            self.lines_per_liaison[key] = []
                        self.lines_per_liaison[key] += value

                    assert self.all_processed_lines
            assert self.all_processed_lines

        with logger_config.stopwatch_with_label("Create all_changement_etats_liaisons_mpro_per_link", inform_beginning=True, monitor_ram_usage=True):
            for changement_etat_liaison_mpro in self.all_changement_etats_liaisons_mpro:
                if changement_etat_liaison_mpro.liaison not in self.all_changement_etats_liaisons_mpro_per_link:
                    self.all_changement_etats_liaisons_mpro_per_link[changement_etat_liaison_mpro.liaison] = []
                self.all_changement_etats_liaisons_mpro_per_link[changement_etat_liaison_mpro.liaison].append(changement_etat_liaison_mpro)

        with logger_config.stopwatch_with_label("Create all_problem_enchainement_numero_protocolaire_per_link", monitor_ram_usage=True):
            for problem_enchainement_numero_protocolaire in self.all_problem_enchainement_numero_protocolaire:
                if problem_enchainement_numero_protocolaire.liaison not in self.all_problem_enchainement_numero_protocolaire_per_link:
                    self.all_problem_enchainement_numero_protocolaire_per_link[problem_enchainement_numero_protocolaire.liaison] = []
                self.all_problem_enchainement_numero_protocolaire_per_link[problem_enchainement_numero_protocolaire.liaison].append(problem_enchainement_numero_protocolaire)

        with logger_config.stopwatch_with_label("Create all_temporary_loss_link", monitor_ram_usage=True, inform_beginning=True):
            for link in self.all_changement_etats_liaisons_mpro_per_link.keys():  # pylint: disable=consider-iterating-dictionary
                last_change_to_nok = None
                for mpro_link_state_change in self.all_changement_etats_liaisons_mpro_per_link[link]:
                    if mpro_link_state_change.new_state == EtatLiaisonMpro.KO:
                        last_change_to_nok = mpro_link_state_change
                    elif mpro_link_state_change.new_state == EtatLiaisonMpro.OK and mpro_link_state_change.previous_state == EtatLiaisonMpro.KO:
                        if last_change_to_nok:
                            assert (
                                last_change_to_nok.trace_line.decoded_timestamp <= mpro_link_state_change.trace_line.decoded_timestamp
                            ), f"Trace to NOK {last_change_to_nok.trace_line.parent_file.file_name}:{last_change_to_nok.trace_line.line_number} ({last_change_to_nok.trace_line.full_raw_line}) is after change to OK {mpro_link_state_change.trace_line.parent_file.file_name}:{mpro_link_state_change.trace_line.line_number} ({mpro_link_state_change.trace_line.full_raw_line})"
                            self.all_temporary_loss_link.append(CckMproTemporaryLossLink(loss_link_event=last_change_to_nok, link_back_to_normal_event=mpro_link_state_change))
                        else:
                            logger_config.print_and_log_error(f"Cannot find change to NOK before {mpro_link_state_change.trace_line.full_raw_line}")

        logger_config.print_and_log_info(f"{len(self.all_problem_enchainement_numero_protocolaire)} problems enchainement numero protocolaire")
        logger_config.print_and_log_info(f"{len(self.all_temporary_loss_link)} temporary_loss_link")
        logger_config.print_and_log_info(f"{len(self.all_changement_etats_liaisons_mpro)} changements états liaisons mpro")
        logger_config.print_and_log_info(
            f"{','.join([key.identifier if key is not None else "None" + ':' + str(len(value)) for key, value in self.lines_per_liaison.items()])} changements états liaisons mpro"
        )
        logger_config.print_and_log_info(f"{len(self.all_problem_enchainement_numero_protocolaire)} all_problem_enchainement_numero_protocolaire")
        logger_config.print_and_log_info(
            f"{','.join([key.identifier if key is not None else "None" + ':' + str(len(value)) for key, value in self.all_problem_enchainement_numero_protocolaire_per_link.items()])} all_problem_enchainement_numero_protocolaire_per_link"
        )

        return self

    def plot_problems_and_loss_link_by_period(self, output_folder_path: str, interval_minutes: int = 10, do_show: bool = False) -> None:
        """
        Génère un bar graph montrant les problèmes d'enchaînement et les pertes de lien par intervalle de temps.

        Args:
            output_folder_path: Chemin du dossier de sortie
            interval_minutes: Intervalle de temps en minutes (par défaut 10)
        """
        if not self.all_processed_lines:
            logger_config.print_and_log_info("La liste des traces est vide. Aucun fichier créé.")
            return

        # Déterminer la période totale
        start_time = self.all_processed_lines[0].decoded_timestamp
        end_time = self.all_processed_lines[-1].decoded_timestamp

        # Créer des intervalles de temps
        interval_start_times: List[datetime.datetime] = []
        current_time = start_time
        while current_time <= end_time:
            interval_start_times.append(current_time)
            current_time += datetime.timedelta(minutes=interval_minutes)

        # Compter les éléments dans chaque intervalle
        interval_problems_count: Dict[Tuple[datetime.datetime, datetime.datetime], int] = {}
        interval_loss_link_count: Dict[Tuple[datetime.datetime, datetime.datetime], int] = {}

        # Initialiser les compteurs
        for interval_start in interval_start_times:
            interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
            interval_problems_count[(interval_start, interval_end)] = 0
            interval_loss_link_count[(interval_start, interval_end)] = 0

        # Compter les problèmes d'enchaînement
        for problem in self.all_problem_enchainement_numero_protocolaire:
            for interval_start in interval_start_times:
                interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                if interval_start <= problem.trace_line.decoded_timestamp < interval_end:
                    interval_problems_count[(interval_start, interval_end)] += 1
                    break

        # Compter les pertes de lien
        for loss_link in self.all_temporary_loss_link:
            for interval_start in interval_start_times:
                interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                if interval_start <= loss_link.loss_link_event.trace_line.decoded_timestamp < interval_end:
                    interval_loss_link_count[(interval_start, interval_end)] += 1
                    break

        # Préparer les données pour le graphe
        x_labels = [f"{begin.strftime("%H:%M")} - {end.strftime("%H:%M")}" for begin, end in interval_problems_count.keys()]
        y_problems = list(interval_problems_count.values())
        y_loss_link = list(interval_loss_link_count.values())

        # Créer et exporter les données dans un fichier Excel
        excel_filename = f"problems_and_loss_link_by_period_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Problems & Loss Link"

        # Ajouter les en-têtes
        ws["A1"] = "Début Intervalle de temps"
        ws["B1"] = "Fin Intervalle de temps"
        ws["C1"] = "Problèmes d'enchaînement"
        ws["D1"] = "Pertes de lien"

        # Style des en-têtes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in [ws["A1"], ws["B1"], ws["C1"], ws["D1"]]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajouter les données
        for idx, ((interval_begin, interval_end), problems_count) in enumerate(interval_problems_count.items(), start=2):
            ws[f"A{idx}"] = interval_begin
            ws[f"B{idx}"] = interval_end
            ws[f"C{idx}"] = problems_count
            ws[f"D{idx}"] = interval_loss_link_count[(interval_begin, interval_end)]
            ws[f"C{idx}"].alignment = Alignment(horizontal="center")
            ws[f"D{idx}"].alignment = Alignment(horizontal="center")

        # Ajuster la largeur des colonnes
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25

        # Ajouter un résumé
        summary_row = len(interval_problems_count) + 3
        ws[f"A{summary_row}"] = "Total"
        ws[f"C{summary_row}"] = len(self.all_problem_enchainement_numero_protocolaire)
        ws[f"D{summary_row}"] = len(self.all_temporary_loss_link)
        ws[f"A{summary_row}"].font = Font(bold=True)
        ws[f"C{summary_row}"].font = Font(bold=True)
        ws[f"D{summary_row}"].font = Font(bold=True)
        ws[f"C{summary_row}"].alignment = Alignment(horizontal="center")
        ws[f"D{summary_row}"].alignment = Alignment(horizontal="center")

        # Sauvegarder le fichier Excel
        wb.save(output_folder_path + "/" + excel_filename)
        logger_config.print_and_log_info(f"Fichier Excel créé: {excel_filename}")

        # Créer et sauvegarder le graphe en HTML avec Plotly
        html_filename = f"problems_and_loss_link_by_period_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.html"
        fig = go.Figure(
            data=[
                go.Bar(
                    name="Problèmes d'enchaînement",
                    x=x_labels,
                    y=y_problems,
                    marker=dict(color="lightcoral"),
                    text=y_problems,
                    textposition="auto",
                ),
                go.Bar(
                    name="Pertes de lien",
                    x=x_labels,
                    y=y_loss_link,
                    marker=dict(color="lightskyblue"),
                    text=y_loss_link,
                    textposition="auto",
                ),
            ]
        )

        fig.update_layout(
            title=f"Problèmes d'enchaînement et Pertes de lien par périodes de {interval_minutes} minutes entre {start_time.strftime('%Y-%m-%d %H:%M')} et {end_time.strftime('%Y-%m-%d %H:%M')}",
            xaxis_title="Intervalles de temps (heure début - heure fin)",
            yaxis_title="Nombre",
            barmode="group",
            hovermode="x unified",
            template="plotly_white",
            height=600,
            width=1200,
        )

        fig.write_html(output_folder_path + "/" + html_filename)
        logger_config.print_and_log_info(f"Fichier HTML créé: {html_filename}")

        # Afficher le bar graph
        plt.figure(figsize=(14, 6))
        x_pos = range(len(x_labels))
        width = 0.35
        plt.bar([p - width / 2 for p in x_pos], y_problems, width, label="Problèmes d'enchaînement", color="lightcoral")
        plt.bar([p + width / 2 for p in x_pos], y_loss_link, width, label="Pertes de lien", color="lightskyblue")
        plt.xlabel("Intervalles de temps (heure début - heure fin)")
        plt.ylabel("Nombre")
        plt.title(f"Problèmes d'enchaînement et Pertes de lien par périodes de {interval_minutes} minutes entre {start_time.strftime("%Y-%m-%d %H:%M")} et {end_time.strftime("%Y-%m-%d %H:%M")}")
        plt.xticks(x_pos, x_labels, rotation=45, ha="right")
        plt.legend()
        plt.tight_layout()
        if do_show:
            plt.show()

    def export_temporary_loss_link_to_excel(self, output_folder_path: str) -> None:
        """
        Exporte les CckMproTemporaryLossLink dans un fichier Excel.

        Args:
            output_folder_path: Chemin du dossier de sortie
            excel_output_file_name_without_extension: Nom du fichier Excel sans extension
        """
        excel_output_file_name_without_extension = self.name + "_temporary_loss_link_" + file_name_utils.get_file_suffix_with_current_datetime()
        if not self.all_temporary_loss_link:
            logger_config.print_and_log_info("Aucune perte de lien temporaire. Aucun fichier créé.")
            return

        # Créer un workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Temporary Loss Link"

        # Ajouter les en-têtes
        headers = [
            "Liaison",
            "Loss Link Event Timestamp",
            "Loss Link Event Line Number",
            "Link Back to Normal Timestamp",
            "Duration (seconds)",
            "Duration (minutes)",
            "Link Back to Normal Line Number",
            "Previous Problem Enchainement Timestamp",
            "Previous Problem Enchainement Line Number",
            "Previous Problem Enchainement Full Line",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajouter les données
        for row_idx, loss_link in enumerate(self.all_temporary_loss_link, start=2):
            loss_timestamp = loss_link.loss_link_event.trace_line.decoded_timestamp
            loss_line_number = loss_link.loss_link_event.trace_line.line_number
            back_to_normal_timestamp = loss_link.link_back_to_normal_event.trace_line.decoded_timestamp
            back_to_normal_line_number = loss_link.link_back_to_normal_event.trace_line.line_number

            # Calculer la durée
            duration_seconds = (back_to_normal_timestamp - loss_timestamp).total_seconds()
            duration_minutes = duration_seconds / 60

            # Trouver le dernier problème d'enchaînement avant la perte sur cette liaison
            previous_problem = None
            previous_problem_timestamp = None
            previous_problem_line_number = None
            previous_problem_full_raw_line = None

            liaison = loss_link.loss_link_event.liaison
            if liaison in self.all_problem_enchainement_numero_protocolaire_per_link:
                problems_on_liaison = self.all_problem_enchainement_numero_protocolaire_per_link[liaison]
                # Chercher les problèmes avant la perte
                problems_before_loss = [p for p in problems_on_liaison if p.trace_line.decoded_timestamp < loss_timestamp]
                if problems_before_loss:
                    # Trouver le dernier
                    previous_problem = max(problems_before_loss, key=lambda p: p.trace_line.decoded_timestamp)
                    previous_problem_timestamp = previous_problem.trace_line.decoded_timestamp
                    previous_problem_line_number = previous_problem.trace_line.line_number
                    previous_problem_full_raw_line = previous_problem.trace_line.full_raw_line

            # Ajouter les données à la ligne
            column_it = custom_iterator.SimpleIntCustomIncrementDecrement(initial_value=1)
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = loss_link.liaison.identifier
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = loss_timestamp
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = loss_line_number
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = back_to_normal_timestamp
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = duration_seconds
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = duration_minutes
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = back_to_normal_line_number
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = previous_problem_timestamp if previous_problem else "N/A"
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = previous_problem_line_number if previous_problem else "N/A"
            ws.cell(row=row_idx, column=column_it.postfix_increment()).value = previous_problem_full_raw_line if previous_problem else "N/A"

        # Ajuster la largeur des colonnes
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 20

        # Sauvegarder le fichier
        wb.save(output_folder_path + "/" + excel_output_file_name_without_extension + "_" + file_name_utils.get_file_suffix_with_current_datetime() + ".xlsx")
        logger_config.print_and_log_info(f"Fichier Excel créé: {excel_output_file_name_without_extension}.xlsx")
        logger_config.print_and_log_info(f"Total de {len(self.all_temporary_loss_link)} pertes de lien sauvegardées")


pass


@dataclass
class CckMproLiaison:
    full_name: str
    identifier: str
    all_lines: List["CckMproTraceLine"] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.hash_computed = hash(self.full_name)

    def __hash__(self) -> int:
        return self.hash_computed


@dataclass
class CckMproTraceFile:
    parent_folder_full_path: str
    file_name: str
    library: "CckMproTraceLibrary"
    all_processed_lines: List["CckMproTraceLine"] = field(default_factory=list)
    all_problem_enchainement_numero_protocolaire: List["CckMproProblemEnchainementNumeroProtocolaire"] = field(default_factory=list)
    all_changement_etats_liaisons_mpro: List["CckMproChangementEtatLiaison"] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.file_full_path = self.parent_folder_full_path + "/" + self.file_name

        self.lines_per_liaison: Dict[Optional[CckMproLiaison], List["CckMproTraceLine"]] = {}
        with logger_config.stopwatch_with_label(f"Open and read CCK Mpro trace file {self.file_full_path}", inform_beginning=True):
            with open(self.file_full_path, mode="r", encoding="ANSI") as file:
                all_raw_lines = file.readlines()
                logger_config.print_and_log_info(to_print_and_log=f"File {self.file_full_path} has {len(all_raw_lines)} lines", do_not_print=True)
                for line_number, line in enumerate(all_raw_lines):
                    if line_number % 100000 == 0:
                        logger_config.print_and_log_info(f"Handle line {self.file_name}:#{line_number}")
                    processed_line = CckMproTraceLine(parent_file=self, full_raw_line=line, line_number=line_number + 1)
                    if processed_line.changement_etat_liaison:
                        self.all_changement_etats_liaisons_mpro.append(processed_line.changement_etat_liaison)
                    if processed_line.problem_enchainement_numero_protocolaire:
                        self.all_problem_enchainement_numero_protocolaire.append(processed_line.problem_enchainement_numero_protocolaire)
                    if processed_line.liaison not in self.lines_per_liaison:
                        self.lines_per_liaison[processed_line.liaison] = []
                    self.lines_per_liaison[processed_line.liaison].append(processed_line)
                    self.all_processed_lines.append(processed_line)

        logger_config.print_and_log_info(f"{self.file_full_path}: {len(self.all_processed_lines)} lines found")
        assert self.all_processed_lines, f"{self.file_full_path} is empty"


@dataclass
class CckMproTraceLine:
    parent_file: CckMproTraceFile
    full_raw_line: str
    line_number: int

    def __post_init__(self) -> None:
        self.raw_date_str = self.full_raw_line[1:23]
        year = int(self.raw_date_str[:4])
        month = int(self.raw_date_str[5:7])
        day = int(self.raw_date_str[8:10])
        hour = int(self.raw_date_str[11:13])
        minute = int(self.raw_date_str[14:16])
        second = int(self.raw_date_str[17:19])
        millisecond = int(self.raw_date_str[20:22]) * 10

        match_liaison_pattern = LIAISON_PATTERN.match(self.full_raw_line)
        self.liaison_full_name = match_liaison_pattern.group("liaison_full_name") if match_liaison_pattern else None
        self.liaison_id = match_liaison_pattern.group("liaison_id") if match_liaison_pattern else None

        self.liaison = self.parent_file.library.get_or_create_liaison(full_name=self.liaison_full_name, identifier=self.liaison_id) if self.liaison_full_name and self.liaison_id else None
        if self.liaison:
            self.liaison.all_lines.append(self)

        self.problem_enchainement_numero_protocolaire = (
            CckMproProblemEnchainementNumeroProtocolaire(self) if "le msg a un problème de 'enchainement numero protocolaire'" in self.full_raw_line else None
        )
        self.changement_etat_liaison = CckMproChangementEtatLiaison(self) if "- changement d'état : " in self.full_raw_line else None

        # self.liaison_name

        # self./
        self.decoded_timestamp = datetime.datetime(year=year, month=month, day=day, hour=hour, minute=minute, second=second, microsecond=millisecond * 1000)
        pass
