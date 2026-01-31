import datetime
import os
import re
from collections import Counter
from dataclasses import dataclass, field
from enum import Enum, auto
from typing import Dict, List, Optional, Self, Tuple, cast

import matplotlib.pyplot as plt
import plotly.graph_objects as go
from common import file_name_utils, string_utils, custom_iterator
from logger import logger_config
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# CONTENT_OF_FIELD_IN_CASE_OF_DECODING_ERROR = "!!! Decoding Error !!!"


class AlarmLineType(Enum):
    EVT_ALA = auto()
    DEB_ALA = auto()
    FIN_ALA = auto()
    OUV_SESSION = auto()
    FERM_SESSION = auto()
    CMD_ESSAIS = auto()
    CMD_CPT_RENDU = auto()
    CSI = auto()


@dataclass
class TerminalTechniqueEquipmentWithAlarms:
    name: str

    def __post_init__(self) -> None:
        self.alarms: List[TerminalTechniqueAlarm] = []


@dataclass
class TerminalTechniqueAlarm:
    raise_line: "TerminalTechniqueArchivesMaintLogLine"
    full_text: str
    alarm_type: AlarmLineType

    def __post_init__(self) -> None:
        self.end_alarm_line: Optional["TerminalTechniqueArchivesMaintLogLine"] = None
        self.equipment_name = self.full_text.split(" ")[0]
        self.equipment = self.raise_line.parent_file.library.get_or_create_equipment_with_alarms(self.equipment_name)
        self.equipment.alarms.append(self)


@dataclass
class TerminalTechniqueCsiAlarm(TerminalTechniqueAlarm):
    pass


@dataclass
class TerminalTechniqueEventAlarm(TerminalTechniqueAlarm):
    pass


@dataclass
class TerminalTechniqueClosableAlarm(TerminalTechniqueAlarm):

    def __post_init__(self) -> None:
        super().__post_init__()
        self.end_alarm_line: Optional["TerminalTechniqueArchivesMaintLogLine"] = None


@dataclass
class TerminalTechniqueSessionAlarm(TerminalTechniqueClosableAlarm):

    def __post_init__(self) -> None:
        super().__post_init__()
        self.session_name = self.raise_line.alarm_full_text


@dataclass
class TerminalTechniqueMccsHAlarm(TerminalTechniqueClosableAlarm):
    pass


@dataclass
class SaatMissingAcknowledgmentTerminalTechniqueAlarm(TerminalTechniqueEventAlarm):
    def __post_init__(self) -> None:
        super().__post_init__()
        self.chaine = self.raise_line.alarm_full_text.split(",")[3]
        self.repetition = int(self.raise_line.alarm_full_text.split(",")[4][0])


@dataclass
class SaharaTerminalTechniqueAlarm(TerminalTechniqueEventAlarm):
    pass


@dataclass
class TerminalTechniqueArchivesMaintLogBackToPast:
    previous_line: "TerminalTechniqueArchivesMaintLogLine"
    next_line: "TerminalTechniqueArchivesMaintLogLine"


@dataclass
class TerminalTechniqueArchivesMaintLibrary:
    name: str

    def __post_init__(self) -> None:
        self.all_processed_lines: List["TerminalTechniqueArchivesMaintLogLine"] = []
        self.all_processed_files: List["TerminalTechniqueArchivesMaintFile"] = []
        self.currently_opened_alarms: List[TerminalTechniqueClosableAlarm] = []
        self.ignored_end_alarms_without_alarm_begin: List[TerminalTechniqueClosableAlarm] = []
        self.sahara_alarms: List[SaharaTerminalTechniqueAlarm] = []
        self.mccs_hs_alarms: List[TerminalTechniqueMccsHAlarm] = []
        self.sessions_alarms: List[TerminalTechniqueSessionAlarm] = []
        self.equipments_with_alarms: List[TerminalTechniqueEquipmentWithAlarms] = []
        self.back_to_past_detected: List[TerminalTechniqueArchivesMaintLogBackToPast] = []

    def load_folder(self, folder_full_path: str) -> Self:

        last_line = None
        for dirpath, dirnames, filenames in os.walk(folder_full_path):
            for file_name in filenames:
                file = TerminalTechniqueArchivesMaintFile(parent_folder_full_path=dirpath, file_name=file_name, library=self, last_line=last_line)
                last_line = file.last_line
                self.all_processed_files.append(file)
                self.all_processed_lines += file.all_processed_lines

                assert self.all_processed_lines
        assert self.all_processed_lines

        logger_config.print_and_log_info(f"{self.name}: ignored_end_alarms_without_alarm_begin:{len(self.ignored_end_alarms_without_alarm_begin)}")
        logger_config.print_and_log_info(f"{self.name}: currently_opened_alarms:{len(self.currently_opened_alarms)}")
        for equipment in self.equipments_with_alarms:
            logger_config.print_and_log_info(f"{self.name}: Equipment {equipment.name} has {len(equipment.alarms)} alarms")
        assert len(self.ignored_end_alarms_without_alarm_begin) < len(self.all_processed_lines), f"{len(self.ignored_end_alarms_without_alarm_begin)} too meny previously opened alarms"
        assert len(self.ignored_end_alarms_without_alarm_begin) < 50000, f"{len(self.ignored_end_alarms_without_alarm_begin)} ignored previously opened alarms"
        assert len(self.currently_opened_alarms) < 1000, f"{len(self.currently_opened_alarms)} opened alarms"

        return self

    def get_or_create_equipment_with_alarms(self, equipment_name: str) -> "TerminalTechniqueEquipmentWithAlarms":
        equipments_found = [equipment for equipment in self.equipments_with_alarms if equipment.name == equipment_name]
        if equipments_found:
            assert len(equipments_found) == 1
            return equipments_found[0]

        equipment = TerminalTechniqueEquipmentWithAlarms(name=equipment_name)
        self.equipments_with_alarms.append(equipment)
        return equipment

    def export_equipments_with_alarms_to_excel(self, output_folder_path: str, equipment_names_to_ignore: List[str]) -> None:
        """
        Exporte tous les équipements et leurs alarmes dans un fichier Excel.

        Args:
            output_folder_path: Chemin du dossier de sortie
            equipment_names_to_ignore: Liste des noms d'équipements à ignorer
            excel_output_file_name_without_extension: Nom du fichier Excel sans extension
        """

        try:
            excel_output_file_name_without_extension: str = self.name + "_all_issues_"
            if not self.equipments_with_alarms:
                logger_config.print_and_log_error("Aucun équipement avec alarmes. Aucun fichier créé.")
                return

            # Créer un workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.name} Equipments Alarms"

            # Ajouter les en-têtes
            headers = [
                "Equipment Name",
                "Type alarm (class name)",
                "Alarm Type",
                "Raise alarm: Timestamp",
                "Raise alarm: File name",
                "Raise alarm: Line number",
                "End alarm (if any): Timestamp",
                "End alarm (if any): File name",
                "End alarm (if any): Line number",
                "Full Text",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajouter les données
            row_idx = 2
            for equipment in self.equipments_with_alarms:
                if equipment.name not in equipment_names_to_ignore:
                    for alarm in equipment.alarms:
                        column_it = custom_iterator.SimpleIntCustomIncrementDecrement(initial_value=1)
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = equipment.name
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = type(alarm).__name__
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.alarm_type.name
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.raise_line.decoded_timestamp
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.raise_line.parent_file.file_name
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.raise_line.line_number
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.end_alarm_line.decoded_timestamp if alarm.end_alarm_line else "NA"
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.end_alarm_line.parent_file.file_name if alarm.end_alarm_line else "NA"
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.end_alarm_line.line_number if alarm.end_alarm_line else "NA"
                        ws.cell(row=row_idx, column=column_it.postfix_increment()).value = alarm.full_text.strip()
                        row_idx += 1

            # Ajuster la largeur des colonnes
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 35
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 15
            ws.column_dimensions["G"].width = 60

            # Sauvegarder le fichier
            wb.save(output_folder_path + "/" + excel_output_file_name_without_extension + "_" + file_name_utils.get_file_suffix_with_current_datetime() + ".xlsx")
            logger_config.print_and_log_info(f"Fichier Excel créé: {excel_output_file_name_without_extension}.xlsx")
            total_alarms = sum(len(equipment.alarms) for equipment in self.equipments_with_alarms)
            logger_config.print_and_log_info(f"Total de {total_alarms} alarmes sauvegardées")

            # Ajouter un onglet pour les SAHARA alarms
            try:

                ws_sahara = wb.create_sheet("SAHARA Alarms")
                headers_sahara = ["Timestamp", "File Name", "Line Number", "Alarm Type", "Full Text"]
                for col_idx, header in enumerate(headers_sahara, start=1):
                    cell = ws_sahara.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for row_idx, sahara_alarm in enumerate(self.sahara_alarms, start=2):
                    ws_sahara.cell(row=row_idx, column=1).value = sahara_alarm.raise_line.decoded_timestamp
                    ws_sahara.cell(row=row_idx, column=2).value = sahara_alarm.raise_line.parent_file.file_name
                    ws_sahara.cell(row=row_idx, column=3).value = sahara_alarm.raise_line.line_number
                    ws_sahara.cell(row=row_idx, column=4).value = sahara_alarm.alarm_type.name
                    ws_sahara.cell(row=row_idx, column=5).value = sahara_alarm.full_text.strip()

                ws_sahara.column_dimensions["A"].width = 25
                ws_sahara.column_dimensions["B"].width = 30
                ws_sahara.column_dimensions["C"].width = 15
                ws_sahara.column_dimensions["D"].width = 15
                ws_sahara.column_dimensions["E"].width = 60

            except Exception as e:
                logger_config.print_and_log_exception(e)

            try:
                # Ajouter un onglet pour les MCCS H alarms
                ws_mccs = wb.create_sheet("MCCS H Alarms")
                headers_mccs = ["Alarm Type", "Duration in seconds", "Timestamp", "File Name", "Line Number", "End Timestamp", "End File Name", "End Line Number", "Full Text"]
                for col_idx, header in enumerate(headers_mccs, start=1):
                    cell = ws_mccs.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for row_idx, mccs_alarm in enumerate(self.mccs_hs_alarms, start=2):
                    column_it = custom_iterator.SimpleIntCustomIncrementDecrement(initial_value=1)
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.alarm_type.name
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = (
                        (mccs_alarm.end_alarm_line.decoded_timestamp - mccs_alarm.raise_line.decoded_timestamp).total_seconds() if mccs_alarm.end_alarm_line else None
                    )
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.raise_line.decoded_timestamp
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.raise_line.parent_file.file_name
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.raise_line.line_number
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.end_alarm_line.decoded_timestamp if mccs_alarm.end_alarm_line else "NA"
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.end_alarm_line.parent_file.file_name if mccs_alarm.end_alarm_line else "NA"
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.end_alarm_line.line_number if mccs_alarm.end_alarm_line else "NA"
                    ws_mccs.cell(row=row_idx, column=column_it.postfix_increment()).value = mccs_alarm.full_text.strip()

                ws_mccs.column_dimensions["A"].width = 25
                ws_mccs.column_dimensions["B"].width = 30
                ws_mccs.column_dimensions["C"].width = 15
                ws_mccs.column_dimensions["D"].width = 15
                ws_mccs.column_dimensions["E"].width = 25
                ws_mccs.column_dimensions["F"].width = 30
                ws_mccs.column_dimensions["G"].width = 15
                ws_mccs.column_dimensions["H"].width = 60

            except Exception as e:
                logger_config.print_and_log_exception(e)

            try:
                # Ajouter un onglet pour les Back to Past events
                ws_btp = wb.create_sheet("Back to Past")
                headers_btp = ["Previous Line Timestamp", "Previous Line Number", "Next Line Timestamp", "Next Line Number", "Duration (seconds)"]
                for col_idx, header in enumerate(headers_btp, start=1):
                    cell = ws_btp.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for row_idx, back_to_past in enumerate(self.back_to_past_detected, start=2):
                    duration = (back_to_past.next_line.decoded_timestamp - back_to_past.previous_line.decoded_timestamp).total_seconds()
                    column_it = custom_iterator.SimpleIntCustomIncrementDecrement(initial_value=1)

                    ws_btp.cell(row=row_idx, column=column_it.postfix_increment()).value = back_to_past.previous_line.decoded_timestamp
                    ws_btp.cell(row=row_idx, column=column_it.postfix_increment()).value = back_to_past.previous_line.line_number
                    ws_btp.cell(row=row_idx, column=column_it.postfix_increment()).value = back_to_past.next_line.decoded_timestamp
                    ws_btp.cell(row=row_idx, column=column_it.postfix_increment()).value = back_to_past.next_line.line_number
                    ws_btp.cell(row=row_idx, column=column_it.postfix_increment()).value = duration

                ws_btp.column_dimensions["A"].width = 25
                ws_btp.column_dimensions["B"].width = 15
                ws_btp.column_dimensions["C"].width = 25
                ws_btp.column_dimensions["D"].width = 15
                ws_btp.column_dimensions["E"].width = 20

            except Exception as e:
                logger_config.print_and_log_exception(e)

            # Sauvegarder le fichier avec tous les onglets
            wb.save(output_folder_path + "/" + excel_output_file_name_without_extension + "_" + file_name_utils.get_file_suffix_with_current_datetime() + ".xlsx")
            logger_config.print_and_log_info(
                f"Onglets supplémentaires créés: SAHARA Alarms ({len(self.sahara_alarms)}), MCCS H Alarms ({len(self.mccs_hs_alarms)}), Back to Past ({len(self.back_to_past_detected)})"
            )
        except Exception as e:
            logger_config.print_and_log_exception(e)

    def plot_alarms_by_period(self, output_folder_path: str, equipment_names_to_ignore: List[str], interval_minutes: int = 10, do_show: bool = False) -> None:
        """
        Génère un bar graph montrant les événements par intervalle de temps.

        Args:
            output_folder_path: Chemin du dossier de sortie
            equipment_names_to_ignore: Liste des noms d'équipements à ignorer
            interval_minutes: Intervalle de temps en minutes (par défaut 10)
            do_show: Afficher le graphique matplotlib
        """
        try:
            if not self.all_processed_lines:
                logger_config.print_and_log_error("La liste des traces est vide. Aucun fichier créé.")
                return

            with logger_config.stopwatch_with_label("plot_alarms_by_period"):

                # Déterminer la période totale
                start_time = self.all_processed_lines[0].decoded_timestamp
                end_time = self.all_processed_lines[-1].decoded_timestamp

                # Créer des intervalles de temps
                interval_start_times: List[datetime.datetime] = []
                current_time = start_time
                while current_time <= end_time:
                    interval_start_times.append(current_time)
                    current_time += datetime.timedelta(minutes=interval_minutes)

                # Initialiser les compteurs pour chaque intervalle
                interval_back_to_past_count: Dict[Tuple[datetime.datetime, datetime.datetime], int] = {}
                interval_sahara_count: Dict[Tuple[datetime.datetime, datetime.datetime], int] = {}
                interval_equipment_counts: Dict[Tuple[datetime.datetime, datetime.datetime], Dict[str, int]] = {}

                for interval_start in interval_start_times:
                    interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                    interval_back_to_past_count[(interval_start, interval_end)] = 0
                    interval_sahara_count[(interval_start, interval_end)] = 0
                    interval_equipment_counts[(interval_start, interval_end)] = {}

                # Compter les back_to_past_detected
                for back_to_past in self.back_to_past_detected:
                    timestamp = back_to_past.previous_line.decoded_timestamp
                    for interval_start in interval_start_times:
                        interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                        if interval_start <= timestamp < interval_end:
                            interval_back_to_past_count[(interval_start, interval_end)] += 1
                            break

                # Compter les sahara_alarms
                for sahara_alarm in self.sahara_alarms:
                    timestamp = sahara_alarm.raise_line.decoded_timestamp
                    for interval_start in interval_start_times:
                        interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                        if interval_start <= timestamp < interval_end:
                            interval_sahara_count[(interval_start, interval_end)] += 1
                            break

                # Compter les alarmes par équipement
                for equipment in self.equipments_with_alarms:
                    if equipment.name not in equipment_names_to_ignore:
                        for alarm in equipment.alarms:
                            timestamp = alarm.raise_line.decoded_timestamp
                            for interval_start in interval_start_times:
                                interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                                if interval_start <= timestamp < interval_end:
                                    if equipment.name not in interval_equipment_counts[(interval_start, interval_end)]:
                                        interval_equipment_counts[(interval_start, interval_end)][equipment.name] = 0
                                    interval_equipment_counts[(interval_start, interval_end)][equipment.name] += 1
                                    break

                # Préparer les données pour le graphe
                x_labels = [f"{begin.strftime("('%Y%m%d_%H:%M")} - {end.strftime("%H:%M")}" for begin, end in interval_back_to_past_count.keys()]
                y_back_to_past = list(interval_back_to_past_count.values())
                y_sahara = list(interval_sahara_count.values())

                # Récupérer tous les noms d'équipements uniques (en ordre d'apparition)
                equipment_names: List[str] = []
                for interval_counts in interval_equipment_counts.values():
                    for name in interval_counts.keys():
                        if name not in equipment_names:
                            equipment_names.append(name)

                # Créer et exporter les données dans un fichier Excel
                excel_filename = f"alarms_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}_{end_time.strftime('%Y%m%d')}_{file_name_utils.get_file_suffix_with_current_datetime()}.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = f"{self.name} Alarms By Period"

                # Ajouter les en-têtes
                headers = ["Début Intervalle", "Fin Intervalle", "Back to Past", "Sahara"]
                headers.extend(equipment_names)

                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Ajouter les données
                for idx, ((interval_begin, interval_end), back_to_past_count) in enumerate(interval_back_to_past_count.items(), start=2):
                    ws[f"A{idx}"] = interval_begin
                    ws[f"B{idx}"] = interval_end
                    ws[f"C{idx}"] = back_to_past_count
                    ws[f"D{idx}"] = interval_sahara_count[(interval_begin, interval_end)]

                    for col_idx, equipment_name in enumerate(equipment_names, start=5):
                        count = interval_equipment_counts[(interval_begin, interval_end)].get(equipment_name, 0)
                        ws.cell(row=idx, column=col_idx).value = count
                        ws.cell(row=idx, column=col_idx).alignment = Alignment(horizontal="center")

                    ws[f"C{idx}"].alignment = Alignment(horizontal="center")
                    ws[f"D{idx}"].alignment = Alignment(horizontal="center")

                # Ajuster la largeur des colonnes
                ws.column_dimensions["A"].width = 25
                ws.column_dimensions["B"].width = 25
                ws.column_dimensions["C"].width = 15
                ws.column_dimensions["D"].width = 15
                for col_idx, _ in enumerate(equipment_names, start=5):
                    col_letter = chr(64 + col_idx)  # Convert number to letter
                    ws.column_dimensions[col_letter].width = 20

                # Ajouter un résumé
                summary_row = len(interval_back_to_past_count) + 3
                ws[f"A{summary_row}"] = "Total"
                ws[f"C{summary_row}"] = len(self.back_to_past_detected)
                ws[f"D{summary_row}"] = len(self.sahara_alarms)
                ws[f"A{summary_row}"].font = Font(bold=True)
                ws[f"C{summary_row}"].font = Font(bold=True)
                ws[f"D{summary_row}"].font = Font(bold=True)
                ws[f"C{summary_row}"].alignment = Alignment(horizontal="center")
                ws[f"D{summary_row}"].alignment = Alignment(horizontal="center")

                # Sauvegarder le fichier Excel
                wb.save(output_folder_path + "/" + excel_filename)
                logger_config.print_and_log_info(f"Fichier Excel créé: {excel_filename}")

                # Créer et sauvegarder le graphe en HTML avec Plotly
                html_filename = f"alarms_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.html"

                fig_data = [
                    go.Bar(
                        name="Back to Past",
                        x=x_labels,
                        y=y_back_to_past,
                        marker=dict(color="orangered"),
                        text=y_back_to_past,
                        textposition="auto",
                    ),
                    go.Bar(
                        name="Sahara",
                        x=x_labels,
                        y=y_sahara,
                        marker=dict(color="gold"),
                        text=y_sahara,
                        textposition="auto",
                    ),
                ]

                # Ajouter les équipements
                colors = ["steelblue", "lightseagreen", "mediumseagreen", "lightcoral", "plum", "khaki", "lightcyan", "lightsalmon"]
                for equipment_idx, equipment_name in enumerate(equipment_names):
                    y_equipment = [
                        interval_equipment_counts[(interval_start, interval_start + datetime.timedelta(minutes=interval_minutes))].get(equipment_name, 0) for interval_start in interval_start_times
                    ]
                    fig_data.append(
                        go.Bar(
                            name=equipment_name,
                            x=x_labels,
                            y=y_equipment,
                            marker=dict(color=colors[equipment_idx % len(colors)]),
                            text=y_equipment,
                            textposition="auto",
                        )
                    )

                fig = go.Figure(data=fig_data)

                fig.update_layout(
                    title=f"{self.name} Alarmes par périodes de {interval_minutes} minutes entre {start_time.strftime('%Y-%m-%d %H:%M')} et {end_time.strftime('%Y-%m-%d %H:%M')}",
                    xaxis_title="Intervalles de temps (heure début - heure fin)",
                    yaxis_title="Nombre",
                    barmode="group",
                    hovermode="x unified",
                    template="plotly_white",
                    height=600,
                    width=1400,
                )

                fig.write_html(output_folder_path + "/" + html_filename)
                logger_config.print_and_log_info(f"Fichier HTML créé: {html_filename}")

                # Afficher le bar graph matplotlib
                plt.figure(figsize=(16, 8))
                x_pos = range(len(x_labels))
                width = 0.8 / (len(equipment_names) + 2)

                plt.bar([p - 0.4 for p in x_pos], y_back_to_past, width, label="Back to Past", color="orangered")
                plt.bar([p - 0.4 + width for p in x_pos], y_sahara, width, label="Sahara", color="gold")

                for equipment_idx, equipment_name in enumerate(equipment_names):
                    y_equipment = [
                        interval_equipment_counts[(interval_start, interval_start + datetime.timedelta(minutes=interval_minutes))].get(equipment_name, 0) for interval_start in interval_start_times
                    ]
                    plt.bar([p - 0.4 + width * (2 + equipment_idx) for p in x_pos], y_equipment, width, label=equipment_name, color=colors[equipment_idx % len(colors)])

                plt.xlabel("Intervalles de temps (heure début - heure fin)")
                plt.ylabel("Nombre")
                plt.title(f"{self.name} Alarmes par périodes de {interval_minutes} minutes entre {start_time.strftime("%Y-%m-%d %H:%M")} et {end_time.strftime("%Y-%m-%d %H:%M")}")
                plt.xticks(x_pos, x_labels, rotation=45, ha="right")
                plt.legend(loc="upper left", bbox_to_anchor=(1, 1))
                plt.tight_layout()
                if do_show:
                    plt.show()

        except Exception as e:
            logger_config.print_and_log_exception(e)

    def plot_sahara_alarms_by_period(self, output_folder_path: str, interval_minutes: int = 10, do_show: bool = False) -> None:
        """
        Génère un bar graph montrant les alarmes SAHARA par intervalle de temps.

        Args:
            output_folder_path: Chemin du dossier de sortie
            interval_minutes: Intervalle de temps en minutes (par défaut 10)
            do_show: Afficher le graphique matplotlib
        """

        try:
            if not self.all_processed_lines:
                logger_config.print_and_log_error("La liste des traces est vide. Aucun fichier créé.")
                return

            # Déterminer la période totale
            start_time = self.all_processed_lines[0].decoded_timestamp
            end_time = self.all_processed_lines[-1].decoded_timestamp

            # Créer des intervalles de temps
            intervals: List[Tuple[datetime.datetime, datetime.datetime]] = []
            interval_start_times: List[datetime.datetime] = []
            current_time = start_time
            while current_time <= end_time:
                interval_start_time = current_time
                interval_start_times.append(current_time)
                current_time += datetime.timedelta(minutes=interval_minutes)
                interval_end_time = current_time
                intervals.append((interval_start_time, interval_end_time))

            logger_config.print_and_log_info(f"plot_sahara_alarms_by_period: {len(interval_start_times)} intervals of {interval_minutes} minutes between {start_time} and {end_time}")

            interval_sahara_counts: Dict[Tuple[datetime.datetime, datetime.datetime], int] = Counter()

            for interval in intervals:
                interval_sahara_counts[interval] = 0

            # Compter les alarmes SAHARA dans chaque intervalle
            for sahara_alarm in self.sahara_alarms:
                timestamp = sahara_alarm.raise_line.decoded_timestamp
                for interval_start in interval_start_times:
                    interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                    if interval_start <= timestamp < interval_end:
                        # logger_config.print_and_log_info(f"plot_sahara_alarms_by_period: Sahara alarm {sahara_alarm.raise_line.full_raw_line} counted in interval {(interval_start, interval_end)}")
                        interval_sahara_counts[(interval_start, interval_end)] += 1
                        break

            # Préparer les données pour le graphe
            x_labels = [f"{begin.strftime("('%Y%m%d %H:%M")} - {end.strftime("%H:%M")}" for begin, end in interval_sahara_counts.keys()]
            y_values = list(interval_sahara_counts.values())

            # Créer et exporter les données dans un fichier Excel
            excel_filename = f"sahara_alarms_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.name} Sahara Alarms"

            # Ajouter les en-têtes
            ws["A1"] = "Début Intervalle de temps"
            ws["B1"] = "Fin Intervalle de temps"
            ws["C1"] = "Nombre d'alarmes SAHARA"

            # Style des en-têtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in [ws["A1"], ws["B1"], ws["C1"]]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajouter les données
            for idx, ((interval_begin, interval_end), count) in enumerate(interval_sahara_counts.items(), start=2):
                ws[f"A{idx}"] = interval_begin
                ws[f"B{idx}"] = interval_end
                ws[f"C{idx}"] = count
                ws[f"C{idx}"].alignment = Alignment(horizontal="center")

            # Ajuster la largeur des colonnes
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 25

            # Ajouter un résumé
            summary_row = len(interval_sahara_counts) + 3
            ws[f"A{summary_row}"] = "Total"
            ws[f"C{summary_row}"] = len(self.sahara_alarms)
            ws[f"A{summary_row}"].font = Font(bold=True)
            ws[f"C{summary_row}"].font = Font(bold=True)
            ws[f"C{summary_row}"].alignment = Alignment(horizontal="center")

            # Sauvegarder le fichier
            wb.save(output_folder_path + "/" + excel_filename)
            logger_config.print_and_log_info(f"Fichier Excel créé: {excel_filename}")

            # Créer et sauvegarder le graphe en HTML avec Plotly
            html_filename = f"sahara_alarms_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.html"
            fig = go.Figure(
                data=[
                    go.Bar(
                        x=x_labels,
                        y=y_values,
                        marker=dict(color="gold"),
                        text=y_values,
                        textposition="auto",
                    )
                ]
            )

            fig.update_layout(
                title=f"{self.name} {len(self.sahara_alarms)} alarmes SAHARA par périodes de {interval_minutes} minutes entre {start_time.strftime('%Y-%m-%d %H:%M')} et {end_time.strftime('%Y-%m-%d %H:%M')}",
                xaxis_title="Intervalles de temps (heure début - heure fin)",
                yaxis_title="Nombre d'alarmes SAHARA",
                hovermode="x unified",
                template="plotly_white",
                height=600,
                width=1000,
            )

            fig.write_html(output_folder_path + "/" + html_filename)
            logger_config.print_and_log_info(f"Fichier HTML créé: {html_filename}")

            # Afficher le bar graph
            plt.figure(figsize=(12, 6))
            plt.bar(x_labels, y_values, color="gold", width=0.6)
            plt.xlabel("Intervalles de temps (heure début - heure fin)")
            plt.ylabel("Nombre d'alarmes SAHARA")
            plt.title(
                f"{self.name} {len(self.sahara_alarms)} alarmes SAHARA par périodes de {interval_minutes} minutes entre {start_time.strftime("%Y-%m-%d %H:%M")} et {end_time.strftime("%Y-%m-%d %H:%M")}"
            )
            plt.xticks(rotation=45, ha="right")
            plt.tight_layout()
            if do_show:
                plt.show()

        except Exception as e:
            logger_config.print_and_log_exception(e)

    def plot_back_to_past_by_period(self, output_folder_path: str, interval_minutes: int = 10, do_show: bool = False) -> None:
        """
        Génère un bar graph montrant les événements back_to_past par intervalle de temps.

        Args:
            output_folder_path: Chemin du dossier de sortie
            interval_minutes: Intervalle de temps en minutes (par défaut 10)
            do_show: Afficher le graphique matplotlib
        """

        try:
            if not self.all_processed_lines:
                logger_config.print_and_log_error("La liste des traces est vide. Aucun fichier créé.")
                return

            # Déterminer la période totale
            start_time = self.all_processed_lines[0].decoded_timestamp
            end_time = self.all_processed_lines[-1].decoded_timestamp

            # Créer des intervalles de temps
            intervals: List[Tuple[datetime.datetime, datetime.datetime]] = []
            interval_start_times: List[datetime.datetime] = []
            current_time = start_time
            while current_time <= end_time:
                interval_start_time = current_time
                interval_start_times.append(current_time)
                current_time += datetime.timedelta(minutes=interval_minutes)
                interval_end_time = current_time
                intervals.append((interval_start_time, interval_end_time))

            interval_back_to_past_counts: Dict[Tuple[datetime.datetime, datetime.datetime], int] = Counter()
            for interval in intervals:
                interval_back_to_past_counts[interval] = 0
            # Compter les back_to_past_detected dans chaque intervalle
            for back_to_past in self.back_to_past_detected:
                timestamp = back_to_past.previous_line.decoded_timestamp
                for interval_start in interval_start_times:
                    interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                    if interval_start <= timestamp < interval_end:
                        interval_back_to_past_counts[(interval_start, interval_end)] += 1
                        break

            # Préparer les données pour le graphe
            x_labels = [f"{begin.strftime("%Y%m%d_%H:%M")} - {end.strftime("%H:%M")}" for begin, end in interval_back_to_past_counts.keys()]
            y_values = list(interval_back_to_past_counts.values())

            # Créer et exporter les données dans un fichier Excel
            excel_filename = f"back_to_past_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.name} Back to Past"

            # Ajouter les en-têtes
            ws["A1"] = "Début Intervalle de temps"
            ws["B1"] = "Fin Intervalle de temps"
            ws["C1"] = "Nombre d'événements Back to Past"

            # Style des en-têtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in [ws["A1"], ws["B1"], ws["C1"]]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajouter les données
            for idx, ((interval_begin, interval_end), count) in enumerate(interval_back_to_past_counts.items(), start=2):
                ws[f"A{idx}"] = interval_begin
                ws[f"B{idx}"] = interval_end
                ws[f"C{idx}"] = count
                ws[f"C{idx}"].alignment = Alignment(horizontal="center")

            # Ajuster la largeur des colonnes
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 30

            # Ajouter un résumé
            summary_row = len(interval_back_to_past_counts) + 3
            ws[f"A{summary_row}"] = "Total"
            ws[f"C{summary_row}"] = len(self.back_to_past_detected)
            ws[f"A{summary_row}"].font = Font(bold=True)
            ws[f"C{summary_row}"].font = Font(bold=True)
            ws[f"C{summary_row}"].alignment = Alignment(horizontal="center")

            # Sauvegarder le fichier
            wb.save(output_folder_path + "/" + excel_filename)
            logger_config.print_and_log_info(f"Fichier Excel créé: {excel_filename}")

            # Créer et sauvegarder le graphe en HTML avec Plotly
            html_filename = f"back_to_past_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.html"
            fig = go.Figure(
                data=[
                    go.Bar(
                        x=x_labels,
                        y=y_values,
                        marker=dict(color="orangered"),
                        text=y_values,
                        textposition="auto",
                    )
                ]
            )

            fig.update_layout(
                title=f"{self.name} {len(self.back_to_past_detected)} événements Back to Past par périodes de {interval_minutes} minutes entre {start_time.strftime('%Y-%m-%d %H:%M')} et {end_time.strftime('%Y-%m-%d %H:%M')}",
                xaxis_title="Intervalles de temps (heure début - heure fin)",
                yaxis_title="Nombre d'événements Back to Past",
                hovermode="x unified",
                template="plotly_white",
                height=600,
                width=1000,
            )

            fig.write_html(output_folder_path + "/" + html_filename)
            logger_config.print_and_log_info(f"Fichier HTML créé: {html_filename}")

            # Afficher le bar graph
            plt.figure(figsize=(12, 6))
            plt.bar(x_labels, y_values, color="orangered", width=0.6)
            plt.xlabel("Intervalles de temps (heure début - heure fin)")
            plt.ylabel("Nombre d'événements Back to Past")
            plt.title(
                f"{self.name} {len(self.back_to_past_detected)} événements Back to Past par périodes de {interval_minutes} minutes entre {start_time.strftime("%Y-%m-%d %H:%M")} et {end_time.strftime("%Y-%m-%d %H:%M")}"
            )
            plt.xticks(rotation=45, ha="right")
            plt.tight_layout()
            if do_show:
                plt.show()

        except Exception as e:
            logger_config.print_and_log_exception(e)

    def plot_sahara_mccs_back_to_past_by_period(self, output_folder_path: str, interval_minutes: int = 10, do_show: bool = False) -> None:
        """
        Génère un bar graph montrant SAHARA, MCCS H et Back to Past par intervalle de temps.

        Args:
            output_folder_path: Chemin du dossier de sortie
            interval_minutes: Intervalle de temps en minutes (par défaut 10)
            do_show: Afficher le graphique matplotlib
        """

        try:
            if not self.all_processed_lines:
                logger_config.print_and_log_error("La liste des traces est vide. Aucun fichier créé.")
                return

            # Déterminer la période totale
            start_time = self.all_processed_lines[0].decoded_timestamp
            end_time = self.all_processed_lines[-1].decoded_timestamp

            # Créer des intervalles de temps
            intervals: List[Tuple[datetime.datetime, datetime.datetime]] = []
            interval_start_times: List[datetime.datetime] = []
            current_time = start_time
            while current_time <= end_time:
                interval_start_time = current_time
                interval_start_times.append(current_time)
                current_time += datetime.timedelta(minutes=interval_minutes)
                interval_end_time = current_time
                intervals.append((interval_start_time, interval_end_time))

            # Compter les événements dans chaque intervalle
            interval_sahara_counts: Dict[Tuple[datetime.datetime, datetime.datetime], int] = Counter()
            interval_mccs_counts: Dict[Tuple[datetime.datetime, datetime.datetime], int] = Counter()
            interval_back_to_past_counts: Dict[Tuple[datetime.datetime, datetime.datetime], int] = Counter()

            for interval in intervals:
                interval_sahara_counts[interval] = 0
                interval_mccs_counts[interval] = 0
                interval_back_to_past_counts[interval] = 0

            # Compter les sahara_alarms
            for sahara_alarm in self.sahara_alarms:
                timestamp = sahara_alarm.raise_line.decoded_timestamp
                for interval_start in interval_start_times:
                    interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                    if interval_start <= timestamp < interval_end:
                        interval_sahara_counts[(interval_start, interval_end)] += 1
                        break

            # Compter les mccs_hs_alarms
            for mccs_alarm in self.mccs_hs_alarms:
                timestamp = mccs_alarm.raise_line.decoded_timestamp
                for interval_start in interval_start_times:
                    interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                    if interval_start <= timestamp < interval_end:
                        interval_mccs_counts[(interval_start, interval_end)] += 1
                        break

            # Compter les back_to_past_detected
            for back_to_past in self.back_to_past_detected:
                timestamp = back_to_past.previous_line.decoded_timestamp
                for interval_start in interval_start_times:
                    interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                    if interval_start <= timestamp < interval_end:
                        interval_back_to_past_counts[(interval_start, interval_end)] += 1
                        break

            # Initialiser les compteurs pour les intervalles manquants
            for interval_start in interval_start_times:
                interval_end = interval_start + datetime.timedelta(minutes=interval_minutes)
                if (interval_start, interval_end) not in interval_sahara_counts:
                    interval_sahara_counts[(interval_start, interval_end)] = 0
                if (interval_start, interval_end) not in interval_mccs_counts:
                    interval_mccs_counts[(interval_start, interval_end)] = 0
                if (interval_start, interval_end) not in interval_back_to_past_counts:
                    interval_back_to_past_counts[(interval_start, interval_end)] = 0

            # Préparer les données pour le graphe
            x_labels = [f"{begin.strftime("('%Y%m%d_%H:%M")} - {end.strftime("%H:%M")}" for begin, end in interval_sahara_counts.keys()]
            y_sahara = [interval_sahara_counts[(begin, end)] for begin, end in interval_sahara_counts.keys()]
            y_mccs = [interval_mccs_counts[(begin, end)] for begin, end in interval_sahara_counts.keys()]
            y_back_to_past = [interval_back_to_past_counts[(begin, end)] for begin, end in interval_sahara_counts.keys()]

            # Créer et exporter les données dans un fichier Excel
            excel_filename = f"sahara_mccs_back_to_past_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.name} Summary Events"

            # Ajouter les en-têtes
            ws["A1"] = "Début Intervalle de temps"
            ws["B1"] = "Fin Intervalle de temps"
            ws["C1"] = "Nombre SAHARA"
            ws["D1"] = "Nombre MCCS H"
            ws["E1"] = "Nombre Back to Past"

            # Style des en-têtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in [ws["A1"], ws["B1"], ws["C1"], ws["D1"], ws["E1"]]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajouter les données
            for idx, ((interval_begin, interval_end), sahara_count) in enumerate(interval_sahara_counts.items(), start=2):
                ws[f"A{idx}"] = interval_begin
                ws[f"B{idx}"] = interval_end
                ws[f"C{idx}"] = sahara_count
                ws[f"D{idx}"] = interval_mccs_counts[(interval_begin, interval_end)]
                ws[f"E{idx}"] = interval_back_to_past_counts[(interval_begin, interval_end)]
                ws[f"C{idx}"].alignment = Alignment(horizontal="center")
                ws[f"D{idx}"].alignment = Alignment(horizontal="center")
                ws[f"E{idx}"].alignment = Alignment(horizontal="center")

            # Ajuster la largeur des colonnes
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 20

            # Ajouter un résumé
            summary_row = len(interval_sahara_counts) + 3
            ws[f"A{summary_row}"] = "Total"
            ws[f"C{summary_row}"] = len(self.sahara_alarms)
            ws[f"D{summary_row}"] = len(self.mccs_hs_alarms)
            ws[f"E{summary_row}"] = len(self.back_to_past_detected)
            ws[f"A{summary_row}"].font = Font(bold=True)
            ws[f"C{summary_row}"].font = Font(bold=True)
            ws[f"D{summary_row}"].font = Font(bold=True)
            ws[f"E{summary_row}"].font = Font(bold=True)
            ws[f"C{summary_row}"].alignment = Alignment(horizontal="center")
            ws[f"D{summary_row}"].alignment = Alignment(horizontal="center")
            ws[f"E{summary_row}"].alignment = Alignment(horizontal="center")

            # Sauvegarder le fichier
            wb.save(output_folder_path + "/" + excel_filename)
            logger_config.print_and_log_info(f"Fichier Excel créé: {excel_filename}")

            # Créer et sauvegarder le graphe en HTML avec Plotly
            html_filename = f"sahara_mccs_back_to_past_by_period_{self.name}_{start_time.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.html"
            fig = go.Figure(
                data=[
                    go.Bar(
                        name="SAHARA",
                        x=x_labels,
                        y=y_sahara,
                        marker=dict(color="gold"),
                        text=y_sahara,
                        textposition="auto",
                    ),
                    go.Bar(
                        name="MCCS H",
                        x=x_labels,
                        y=y_mccs,
                        marker=dict(color="steelblue"),
                        text=y_mccs,
                        textposition="auto",
                    ),
                    go.Bar(
                        name="Back to Past",
                        x=x_labels,
                        y=y_back_to_past,
                        marker=dict(color="orangered"),
                        text=y_back_to_past,
                        textposition="auto",
                    ),
                ]
            )

            fig.update_layout(
                title=f"{self.name} SAHARA, MCCS H et Back to Past par périodes de {interval_minutes} minutes entre {start_time.strftime('%Y-%m-%d %H:%M')} et {end_time.strftime('%Y-%m-%d %H:%M')}",
                xaxis_title="Intervalles de temps (heure début - heure fin)",
                yaxis_title="Nombre d'événements",
                barmode="group",
                hovermode="x unified",
                template="plotly_white",
                height=600,
                width=1200,
            )

            fig.write_html(output_folder_path + "/" + html_filename)
            logger_config.print_and_log_info(f"Fichier HTML créé: {html_filename}")

            # Afficher le bar graph matplotlib
            plt.figure(figsize=(14, 7))
            x_pos = range(len(x_labels))
            width = 0.25

            plt.bar([p - width for p in x_pos], y_sahara, width, label="SAHARA", color="gold")
            plt.bar(x_pos, y_mccs, width, label="MCCS H", color="steelblue")
            plt.bar([p + width for p in x_pos], y_back_to_past, width, label="Back to Past", color="orangered")

            plt.xlabel("Intervalles de temps (heure début - heure fin)")
            plt.ylabel("Nombre d'événements")
            plt.title(f"{self.name} SAHARA, MCCS H et Back to Past par périodes de {interval_minutes} minutes entre {start_time.strftime("%Y-%m-%d %H:%M")} et {end_time.strftime("%Y-%m-%d %H:%M")}")
            plt.xticks(x_pos, x_labels, rotation=45, ha="right")
            plt.legend()
            plt.tight_layout()
            if do_show:
                plt.show()

        except Exception as e:
            logger_config.print_and_log_exception(e)

    def dump_all_events_to_text_file(self, output_folder_path: str) -> None:
        """
        Dump tous les événements (MCCS H alarms, SAHARA alarms, Back to Past) 
        dans un fichier texte par ordre chronologique.

        Args:
            output_folder_path: Chemin du dossier de sortie
        """
        try:
            if not self.all_processed_lines:
                logger_config.print_and_log_error("La liste des traces est vide. Aucun fichier créé.")
                return

            # Créer une liste contenant tous les événements avec leurs timestamps
            events: List[Tuple[datetime.datetime, str, str]] = []

            # Ajouter les MCCS H alarms
            for mccs_alarm in self.mccs_hs_alarms:
                timestamp = mccs_alarm.raise_line.decoded_timestamp
                event_text = f"[MCCS H ALARM] {timestamp} | File: {mccs_alarm.raise_line.parent_file.file_name}:{mccs_alarm.raise_line.line_number} | {mccs_alarm.full_text.strip()}"
                events.append((timestamp, "MCCS_H", event_text))

            # Ajouter les SAHARA alarms
            for sahara_alarm in self.sahara_alarms:
                timestamp = sahara_alarm.raise_line.decoded_timestamp
                event_text = f"[SAHARA ALARM] {timestamp} | File: {sahara_alarm.raise_line.parent_file.file_name}:{sahara_alarm.raise_line.line_number} | {sahara_alarm.full_text.strip()}"
                events.append((timestamp, "SAHARA", event_text))

            # Ajouter les Back to Past events
            for back_to_past in self.back_to_past_detected:
                timestamp = back_to_past.previous_line.decoded_timestamp
                next_timestamp = back_to_past.next_line.decoded_timestamp
                time_diff = (next_timestamp - timestamp).total_seconds()
                event_text = f"[BACK TO PAST] {timestamp} | Previous: {back_to_past.previous_line.parent_file.file_name}:{back_to_past.previous_line.line_number} | Next: {back_to_past.next_line.parent_file.file_name}:{back_to_past.next_line.line_number} | Jump: {time_diff:.2f}s backward"
                events.append((timestamp, "BACK_TO_PAST", event_text))

            # Trier tous les événements par timestamp
            events.sort(key=lambda x: x[0])

            # Écrire dans le fichier texte
            text_filename = f"all_events_{self.name}_{self.all_processed_lines[0].decoded_timestamp.strftime('%Y%m%d_%H%M%S')}{file_name_utils.get_file_suffix_with_current_datetime()}.txt"
            
            with open(output_folder_path + "/" + text_filename, mode="w", encoding="utf-8") as f:
                f.write(f"{'='*120}\n")
                f.write(f"Chronological dump of all events for {self.name}\n")
                f.write(f"Total events: {len(events)} (MCCS H: {len(self.mccs_hs_alarms)}, SAHARA: {len(self.sahara_alarms)}, Back to Past: {len(self.back_to_past_detected)})\n")
                f.write(f"{'='*120}\n\n")

                for timestamp, event_type, event_text in events:
                    f.write(event_text + "\n")

            logger_config.print_and_log_info(f"Fichier texte créé: {text_filename}")
            logger_config.print_and_log_info(f"Total de {len(events)} événements exportés (MCCS H: {len(self.mccs_hs_alarms)}, SAHARA: {len(self.sahara_alarms)}, Back to Past: {len(self.back_to_past_detected)})")

        except Exception as e:
            logger_config.print_and_log_exception(e)


@dataclass
class TerminalTechniqueArchivesMaintFile:
    parent_folder_full_path: str
    file_name: str
    library: "TerminalTechniqueArchivesMaintLibrary"
    last_line: Optional["TerminalTechniqueArchivesMaintLogLine"]
    all_processed_lines: List["TerminalTechniqueArchivesMaintLogLine"] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.file_full_path = self.parent_folder_full_path + "\\" + self.file_name

        with logger_config.stopwatch_with_label(f"Open and read file  {self.file_full_path}", inform_beginning=False, enable_print=False, enabled=False):
            with open(self.file_full_path, mode="r", encoding="ANSI") as file:
                all_raw_lines = file.readlines()
                # logger_config.print_and_log_info(to_print_and_log=f"File {self.file_full_path} has {len(all_raw_lines)} lines")
                for line_number, line in enumerate(all_raw_lines):
                    if len(line) > 3:
                        try:
                            processed_line = TerminalTechniqueArchivesMaintLogLine(parent_file=self, full_raw_line=line, line_number=line_number + 1, previous_line=self.last_line)
                            self.all_processed_lines.append(processed_line)
                            self.last_line = processed_line
                        except Exception as e:
                            logger_config.print_and_log_exception(e)
                            logger_config.print_and_log_error(f"Could not process line {line} in {self.file_name} {line_number+1}")


@dataclass
class TerminalTechniqueArchivesMaintLogLine:
    parent_file: TerminalTechniqueArchivesMaintFile
    full_raw_line: str
    line_number: int
    previous_line: Optional["TerminalTechniqueArchivesMaintLogLine"]

    def __post_init__(self) -> None:
        self.raw_date_str = self.full_raw_line[0:22]

        self.full_raw_line_split_by_tab = self.full_raw_line.split("\t")
        self.alarm_type = AlarmLineType[self.full_raw_line_split_by_tab[2]]

        # 2025-12-29	01:45:53.30
        self.raw_date_str_with_microseconds = self.raw_date_str + "0000"
        self.decoded_timestamp = datetime.datetime.strptime(self.raw_date_str_with_microseconds, "%Y-%m-%d	%H:%M:%S.%f")

        if self.previous_line and self.previous_line.decoded_timestamp > self.decoded_timestamp:
            back_to_past = TerminalTechniqueArchivesMaintLogBackToPast(previous_line=self.previous_line, next_line=self)
            self.parent_file.library.back_to_past_detected.append(back_to_past)

        assert len(self.full_raw_line_split_by_tab) == 4
        self.alarm_full_text = self.full_raw_line_split_by_tab[3]

        self.alarm: Optional[TerminalTechniqueAlarm] = None

        if self.alarm_type in [AlarmLineType.FIN_ALA, AlarmLineType.FERM_SESSION]:
            found_unclosed_alarms = [alarm for alarm in self.parent_file.library.currently_opened_alarms if alarm.full_text == self.alarm_full_text]
            if found_unclosed_alarms:
                while len(found_unclosed_alarms) > 1:
                    never_closed_alarm = found_unclosed_alarms[0]
                    logger_config.print_and_log_warning(
                        f"Alarm {never_closed_alarm.full_text} is re-opened at {found_unclosed_alarms[0].raise_line.decoded_timestamp} and not closed between. {len(found_unclosed_alarms)} unclosed alarms found \n({'\n'.join(alarm.raise_line.full_raw_line + " in " + alarm.raise_line.parent_file.file_name + ":"+str(alarm.raise_line.line_number)  for alarm in found_unclosed_alarms)}) for {self.alarm_full_text} when trying to close {self.full_raw_line} in {self.parent_file.file_name}:{self.line_number}"
                    )
                    self.parent_file.library.currently_opened_alarms.remove(found_unclosed_alarms[0])
                    found_unclosed_alarms = [alarm for alarm in self.parent_file.library.currently_opened_alarms if alarm.full_text == self.alarm_full_text]

                found_unclosed_alarm = found_unclosed_alarms[0]
                self.parent_file.library.currently_opened_alarms.remove(found_unclosed_alarm)
                assert found_unclosed_alarm.end_alarm_line is None
                self.alarm = found_unclosed_alarm
                self.alarm.end_alarm_line = self
            else:
                self.alarm = TerminalTechniqueClosableAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
                self.alarm.end_alarm_line = self
                self.parent_file.library.ignored_end_alarms_without_alarm_begin.append(self.alarm)

        elif "SAHARA" in self.alarm_full_text:
            self.alarm = SaharaTerminalTechniqueAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
            self.parent_file.library.sahara_alarms.append(self.alarm)
        elif "Absence acquittement SAAT" in self.alarm_full_text:
            self.alarm = SaatMissingAcknowledgmentTerminalTechniqueAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
        elif self.alarm_type == AlarmLineType.DEB_ALA and ("MCCS A HS" in self.alarm_full_text or "MCCS B HS" in self.alarm_full_text):
            self.alarm = TerminalTechniqueMccsHAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
            self.parent_file.library.mccs_hs_alarms.append(self.alarm)

            self.parent_file.library.currently_opened_alarms.append(self.alarm)
        elif self.alarm_type == AlarmLineType.DEB_ALA:
            self.alarm = TerminalTechniqueClosableAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
            self.parent_file.library.currently_opened_alarms.append(self.alarm)
        elif self.alarm_type == AlarmLineType.OUV_SESSION:
            session_alarm = TerminalTechniqueSessionAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
            self.alarm = session_alarm
            self.parent_file.library.currently_opened_alarms.append(self.alarm)
            self.parent_file.library.sessions_alarms.append(session_alarm)

        elif self.alarm_type in [AlarmLineType.EVT_ALA, AlarmLineType.CMD_ESSAIS, AlarmLineType.CMD_CPT_RENDU]:
            self.alarm = TerminalTechniqueAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
        elif self.alarm_type == AlarmLineType.CSI:
            self.alarm = TerminalTechniqueCsiAlarm(raise_line=self, full_text=self.alarm_full_text, alarm_type=self.alarm_type)
        else:
            assert False, f"Alarm type {self.alarm_type} for {self.parent_file.file_name} {self.line_number} {self.alarm_full_text}"
